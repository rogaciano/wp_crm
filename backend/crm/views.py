"""
Views da API do CRM
"""
import logging
import re
import unicodedata
from datetime import timedelta, timezone as dt_timezone
from decimal import Decimal, InvalidOperation
from rest_framework import viewsets, status, filters, permissions
from rest_framework.views import APIView
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from rest_framework.throttling import AnonRateThrottle
from django.conf import settings
from django.db import transaction
from django.db.models import Q, Sum, Count, Avg, Exists, OuterRef
from django.contrib.auth.models import Permission
from django_filters.rest_framework import DjangoFilterBackend


class DiagnosticoPublicoThrottle(AnonRateThrottle):
    scope = 'diagnostico_publico'


class WebhookThrottle(AnonRateThrottle):
    scope = 'webhook'

logger = logging.getLogger(__name__)


def parse_bool_param(raw_value):
    """Converte query param de boolean para True/False/None."""
    if raw_value is None:
        return None

    value = str(raw_value).strip().lower()
    if value in {'1', 'true', 't', 'yes', 'sim'}:
        return True
    if value in {'0', 'false', 'f', 'no', 'nao', 'não'}:
        return False
    return None

from .models import (
    Canal, User, Conta, Contato, TipoContato, TipoRedeSocial, Funil, EstagioFunil, FunilEstagio, Oportunidade, OportunidadeAnexo, Atividade, Origem,
    DiagnosticoPilar, DiagnosticoPergunta, DiagnosticoResposta, DiagnosticoResultado,
    Plano, PlanoAdicional, WhatsappMessage, Log, NumeroBloqueado,
    ModuloTreinamento, OnboardingCliente, SessaoTreinamento, AgendaTreinamento
)
from .serializers import (
    CanalSerializer, UserSerializer, ContaSerializer, OrigemSerializer,
    ContatoSerializer, TipoContatoSerializer, TipoRedeSocialSerializer, EstagioFunilSerializer, FunilEstagioSerializer, OportunidadeSerializer,
    OportunidadeKanbanSerializer, AtividadeSerializer, OportunidadeAnexoSerializer,
    DiagnosticoPilarSerializer, DiagnosticoResultadoSerializer, DiagnosticoPublicSubmissionSerializer,
    PlanoSerializer, PlanoAdicionalSerializer, FunilSerializer, WhatsappMessageSerializer,
    WhatsappMessageSlimSerializer, LogSerializer,
    TagSerializer, ModuloTreinamentoSerializer,
    OnboardingClienteSerializer, OnboardingClienteListSerializer, SessaoTreinamentoSerializer,
    AgendaTreinamentoSerializer
)
from .services.ai_service import gerar_analise_diagnostico
from .services.evolution_api import EvolutionService
from .permissions import HierarchyPermission, IsAdminUser
from django.utils import timezone
from django.conf import settings


class StandardResultsSetPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 1000


class ChatMessagesPagination(PageNumberPagination):
    """Paginação otimizada para histórico de chat."""
    page_size = 50
    page_size_query_param = 'page_size'
    max_page_size = 200


class CanalViewSet(viewsets.ModelViewSet):
    """ViewSet para Canais (CRUD apenas Admin, leitura para autenticados)"""
    serializer_class = CanalSerializer
    
    def get_permissions(self):
        # Ações de leitura e WhatsApp permitidas para usuários autenticados
        # (get_queryset já filtra para mostrar apenas o canal do usuário)
        whatsapp_actions = [
            'conectar_whatsapp', 'whatsapp_status', 'whatsapp_qrcode',
            'whatsapp_desconectar', 'whatsapp_reiniciar', 'whatsapp_deletar_instancia'
        ]
        if self.action in ['list', 'retrieve'] + whatsapp_actions:
            return [permissions.IsAuthenticated()]
        return [IsAdminUser()]

    def get_queryset(self):
        user = self.request.user
        if user.perfil == 'ADMIN':
            return Canal.objects.all()
        
        # Responsável e Vendedor veem apenas o canal que pertencem 
        # ou que gerenciam (no caso do responsável)
        from django.db.models import Q
        q_filter = Q()
        if user.canal_id:
            q_filter |= Q(id=user.canal_id)
        
        if user.perfil == 'RESPONSAVEL':
            q_filter |= Q(responsavel=user)
            
        return Canal.objects.filter(q_filter).distinct()

    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['nome']
    ordering_fields = ['nome', 'data_criacao']

    # ==================== ENDPOINTS WHATSAPP EVOLUTION ====================
    
    @action(detail=True, methods=['post'], url_path='conectar-whatsapp')
    def conectar_whatsapp(self, request, pk=None):
        """
        Cria/conecta uma instância WhatsApp Evolution para este canal.
        Usa a Global API Key para criar a instância e salva o token retornado.
        """
        canal = self.get_object()
        
        # Se já tem instância, retorna erro
        if canal.evolution_instance_name and canal.evolution_token:
            return Response({
                'success': False,
                'error': 'Este canal já possui uma instância WhatsApp configurada. Desconecte primeiro.'
            }, status=400)
        
        # Gera nome único para a instância baseado no canal
        import re
        instance_name = re.sub(r'[^a-z0-9]', '_', canal.nome.lower())
        instance_name = re.sub(r'_+', '_', instance_name).strip('_')
        instance_name = f"canal_{instance_name}"
        
        # Webhook URL para receber eventos
        webhook_url = request.build_absolute_uri('/api/webhook/whatsapp/')
        
        # Cria a instância usando Global API Key
        evolution = EvolutionService()
        result = evolution.create_instance(instance_name, webhook_url)
        
        if result['success']:
            # Salva os dados retornados
            canal.evolution_instance_name = instance_name
            canal.evolution_token = result.get('token')
            canal.evolution_is_connected = False
            canal.evolution_last_status = 'created'
            canal.save()
            
            return Response({
                'success': True,
                'instance_name': instance_name,
                'token': result.get('token'),
                'qr_code': result.get('qr_code'),
                'qr_base64': result.get('qr_base64'),
                'message': 'Instância criada com sucesso! Escaneie o QR Code para conectar.'
            })
        else:
            return Response({
                'success': False,
                'error': result.get('error', 'Erro ao criar instância')
            }, status=500)

    @action(detail=True, methods=['get'], url_path='whatsapp/status')
    def whatsapp_status(self, request, pk=None):
        """Retorna o status da conexão WhatsApp do canal"""
        canal = self.get_object()
        
        if not canal.evolution_instance_name:
            return Response({
                'connected': False,
                'state': 'not_configured',
                'has_instance': False,
                'message': 'Instância WhatsApp não configurada para este canal'
            })
        
        # Usa o token da instância para verificar status
        evolution = EvolutionService(canal.evolution_instance_name, canal.evolution_token)
        status = evolution.get_connection_status()
        
        # Atualiza o status no banco
        new_connected = status.get('connected', False)
        new_status = status.get('state', 'unknown')
        
        if new_connected != canal.evolution_is_connected or new_status != canal.evolution_last_status:
            canal.evolution_is_connected = new_connected
            canal.evolution_last_status = new_status
            canal.save()
        
        # Se a instância não existe na Evolution API, limpa o registro do banco
        instance_exists = new_status != 'not_found'
        if not instance_exists:
            canal.evolution_instance_name = None
            canal.evolution_token = None
            canal.evolution_is_connected = False
            canal.evolution_last_status = None
            canal.save()
            return Response({
                'connected': False,
                'state': 'not_found',
                'has_instance': False,
                'message': 'A instância não foi encontrada na Evolution API. Reconecte o WhatsApp.'
            })

        return Response({
            'connected': new_connected,
            'state': new_status,
            'has_instance': True,
            'instance_name': canal.evolution_instance_name,
            **status
        })

    @action(detail=True, methods=['get'], url_path='whatsapp/qrcode')
    def whatsapp_qrcode(self, request, pk=None):
        """Retorna o QR Code para conectar o WhatsApp do canal"""
        canal = self.get_object()
        
        if not canal.evolution_instance_name:
            return Response({
                'success': False,
                'error': 'Instância WhatsApp não configurada para este canal'
            }, status=400)
        
        evolution = EvolutionService(canal.evolution_instance_name, canal.evolution_token)
        result = evolution.get_qr_code()
        
        return Response(result)

    @action(detail=True, methods=['post'], url_path='whatsapp/desconectar')
    def whatsapp_desconectar(self, request, pk=None):
        """Desconecta o WhatsApp do canal (logout)"""
        canal = self.get_object()
        
        if not canal.evolution_instance_name:
            return Response({
                'success': False,
                'error': 'Instância WhatsApp não configurada para este canal'
            }, status=400)
        
        evolution = EvolutionService(canal.evolution_instance_name, canal.evolution_token)
        result = evolution.disconnect()
        
        if result.get('success'):
            canal.evolution_is_connected = False
            canal.evolution_last_status = 'disconnected'
            canal.evolution_phone_number = None
            canal.save()
        
        return Response(result)

    @action(detail=True, methods=['post'], url_path='whatsapp/reiniciar')
    def whatsapp_reiniciar(self, request, pk=None):
        """Reinicia a instância WhatsApp do canal"""
        canal = self.get_object()
        
        if not canal.evolution_instance_name:
            return Response({
                'success': False,
                'error': 'Instância WhatsApp não configurada para este canal'
            }, status=400)
        
        evolution = EvolutionService(canal.evolution_instance_name, canal.evolution_token)
        result = evolution.restart_instance()
        
        if result.get('success'):
            canal.evolution_last_status = 'restarting'
            canal.save()
        
        return Response(result)

    @action(detail=True, methods=['delete'], url_path='whatsapp/deletar-instancia')
    def whatsapp_deletar_instancia(self, request, pk=None):
        """Deleta a instância WhatsApp do canal completamente"""
        canal = self.get_object()
        
        if not canal.evolution_instance_name:
            return Response({
                'success': False,
                'error': 'Instância WhatsApp não configurada para este canal'
            }, status=400)
        
        # Usa Global API Key para deletar
        evolution = EvolutionService(canal.evolution_instance_name)
        result = evolution.delete_instance()
        
        if result.get('success'):
            canal.evolution_instance_name = None
            canal.evolution_token = None
            canal.evolution_is_connected = False
            canal.evolution_last_status = None
            canal.evolution_phone_number = None
            canal.save()
        
        return Response(result)


class OrigemViewSet(viewsets.ModelViewSet):
    """ViewSet para Origens (Fontes de Oportunidades)"""
    queryset = Origem.objects.filter(ativo=True)
    serializer_class = OrigemSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['nome']
    ordering_fields = ['nome']

    def get_queryset(self):
        # Show all for reading, but admin can manage all including inactive
        if self.request.user.perfil == 'ADMIN':
            return Origem.objects.all()
        return Origem.objects.filter(ativo=True)

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [permissions.IsAuthenticated()]
        return [IsAdminUser()]


class UserViewSet(viewsets.ModelViewSet):
    """ViewSet para Usuários"""
    queryset = User.objects.all()
    serializer_class = UserSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['perfil', 'canal', 'is_active']
    search_fields = ['username', 'email', 'first_name', 'last_name']
    ordering_fields = ['username', 'date_joined']

    def get_queryset(self):
        qs = super().get_queryset()
        funis_tipo = self.request.query_params.get('funis_tipo')
        if funis_tipo:
            qs = qs.filter(funis_acesso__tipo=funis_tipo).distinct()
        return qs

    def get_permissions(self):
        # Permitir listagem e 'me' para usuários autenticados
        if self.action in ['list', 'retrieve', 'me']:
            return [permissions.IsAuthenticated()]
        return [IsAdminUser()]

    @action(detail=False, methods=['get'])
    def me(self, request):
        """Retorna informações do usuário logado"""
        serializer = self.get_serializer(request.user)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='auth-permissions')
    def auth_permissions(self, request):
        """Lista permissões de auth para gestão (view/add/change/delete por modelo)."""
        if request.user.perfil != 'ADMIN':
            return Response({'detail': 'Sem permissão.'}, status=status.HTTP_403_FORBIDDEN)

        allowed_prefixes = ('view_', 'add_', 'change_', 'delete_')
        permissions_qs = Permission.objects.select_related('content_type').filter(
            codename__startswith=allowed_prefixes
        ).order_by('content_type__app_label', 'content_type__model', 'codename')

        data = [
            {
                'id': perm.id,
                'codename': perm.codename,
                'name': perm.name,
                'app_label': perm.content_type.app_label,
                'model': perm.content_type.model,
            }
            for perm in permissions_qs
        ]
        return Response(data)


class FunilViewSet(viewsets.ModelViewSet):
    """ViewSet para Funis"""
    serializer_class = FunilSerializer
    
    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsAdminUser()]
        return [permissions.IsAuthenticated()]
    
    def get_queryset(self):
        user = self.request.user
        tipo = self.request.query_params.get('tipo')

        # Admin vê todos para gestão
        if user.perfil == 'ADMIN':
            qs = Funil.objects.all()
        else:
            # Para outros perfis, busca funis vinculados ao usuário OU ao canal do usuário
            from django.db.models import Q

            # Funis que o usuário tem acesso direto
            funis_acesso_ids = list(user.funis_acesso.filter(is_active=True).values_list('id', flat=True))

            # Funis vinculados a usuários do mesmo canal
            if user.canal:
                funis_canal_ids = list(Funil.objects.filter(
                    usuarios__canal=user.canal,
                    is_active=True
                ).values_list('id', flat=True))
            else:
                funis_canal_ids = []

            # Combina os dois conjuntos
            # Busca funis vinculados ao usuário OU funis sem restrição (globais)
            all_funil_ids = list(set(funis_acesso_ids + funis_canal_ids))

            from django.db.models import Q
            q_filter = Q(usuarios__isnull=True)
            if all_funil_ids:
                q_filter |= Q(id__in=all_funil_ids)

            qs = Funil.objects.filter(q_filter, is_active=True).distinct()

        # Filtra por tipo se especificado
        if tipo:
            qs = qs.filter(tipo=tipo)

        return qs

    @action(detail=True, methods=['get'])
    def estagios(self, request, pk=None):
        """Lista estágios de um funil específico via tabela de ligação"""
        funil = self.get_object()
        vinculos = FunilEstagio.objects.filter(funil=funil).order_by('ordem')
        serializer = FunilEstagioSerializer(vinculos, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def atualizar_estagios(self, request, pk=None):
        """Atualiza a lista de estágios e sua ordem para o funil"""
        funil = self.get_object()
        # Lista de dicionários [{'estagio_id': 1, 'ordem': 0, 'is_padrao': True}, ...]
        estagios_data = request.data.get('estagios', [])
        
        try:
            with transaction.atomic():
                # Remove vínculos atuais? 
                # (Melhor remover e recriar para garantir a ordem e seleção exata do usuário)
                FunilEstagio.objects.filter(funil=funil).delete()
                
                for item in estagios_data:
                    FunilEstagio.objects.create(
                        funil=funil,
                        estagio_id=item['estagio_id'],
                        ordem=item.get('ordem', 0),
                        is_padrao=item.get('is_padrao', False)
                    )
            
            return Response({'status': 'estágios atualizados'})
        except Exception as e:
            return Response({'error': str(e)}, status=400)


        queryset = self.get_queryset()
        
        # Filtros de busca se enviados
        search = request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(nome__icontains=search) | 
                Q(email__icontains=search) | 
                Q(empresa__icontains=search)
            )
            
        status_filter = request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
            
        fonte_filter = request.query_params.get('fonte')
        if fonte_filter:
            queryset = queryset.filter(fonte=fonte_filter)

        total = queryset.count()
        leads_novos = queryset.filter(status=Lead.STATUS_NOVO).count()
        leads_qualificados = queryset.filter(status=Lead.STATUS_QUALIFICADO).count()
        leads_convertidos = queryset.filter(status=Lead.STATUS_CONVERTIDO).count()
        
        taxa_conversao = (leads_convertidos / total * 100) if total > 0 else 0
        
        return Response({
            'total': total,
            'novos': leads_novos,
            'qualificados': leads_qualificados,
            'convertidos': leads_convertidos,
            'taxa_conversao': round(taxa_conversao, 1)
        })


class ContaViewSet(viewsets.ModelViewSet):
    """ViewSet para Contas"""
    serializer_class = ContaSerializer
    
    @action(detail=False, methods=['get'], url_path='checar_cnpj')
    def checar_cnpj(self, request):
        cnpj = request.query_params.get('cnpj', '')
        cnpj_limpo = ''.join(filter(str.isdigit, cnpj))
        
        if not cnpj_limpo:
            return Response({'exists': False})
            
        from django.db.models import Q
        
        # 1. Busca exata (com máscara ou sem máscara)
        conta = self.get_queryset().filter(Q(cnpj=cnpj) | Q(cnpj=cnpj_limpo)).first()
        
        # 2. Se não achou, filtra em memória de forma segura para não depender do Replace do DB
        if not conta:
            # Pega registros que contêm a raiz do CNPJ (ex: primeiros 8 dígitos)
            qs = self.get_queryset().filter(cnpj__icontains=cnpj_limpo[:8])
            for c in qs:
                c_limpo = ''.join(filter(str.isdigit, c.cnpj or ''))
                if c_limpo == cnpj_limpo:
                    conta = c
                    break
        
        if conta:
            from django.contrib.contenttypes.models import ContentType
            ct = ContentType.objects.get_for_model(conta)
            return Response({
                'exists': True,
                'conta': {
                    'id': conta.id,
                    'nome_empresa': conta.nome_empresa,
                    'content_type_id': ct.id
                }
            })
            
        return Response({'exists': False})
        
    permission_classes = [HierarchyPermission]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['setor', 'estado', 'tags', 'status_cliente']
    search_fields = ['nome_empresa', 'cnpj', 'email']
    ordering_fields = ['nome_empresa', 'data_criacao']
    
    def get_queryset(self):
        user = self.request.user
        
        if user.perfil == 'ADMIN':
            queryset = Conta.objects.all()
        elif user.perfil == 'RESPONSAVEL':
            queryset = Conta.objects.filter(
                Q(proprietario__canal=user.canal) | Q(canal=user.canal)
            ).distinct()
        else: # VENDEDOR
            # Vê do seu canal
            queryset = Conta.objects.filter(
                Q(proprietario__canal=user.canal) | Q(canal=user.canal)
            ).distinct()

        opps_abertas = Oportunidade.objects.filter(
            Q(conta_id=OuterRef('pk')) | Q(empresas__id=OuterRef('pk')),
            estagio__tipo=EstagioFunil.TIPO_ABERTO,
        )
        opps_abertas_vendas = opps_abertas.filter(funil__tipo=Funil.TIPO_VENDAS)

        queryset = queryset.annotate(
            tem_oportunidade_aberta=Exists(opps_abertas),
            tem_oportunidade_upgrade=Exists(opps_abertas_vendas),
        )

        tem_oportunidade_aberta = parse_bool_param(self.request.query_params.get('tem_oportunidade_aberta'))
        if tem_oportunidade_aberta is True:
            queryset = queryset.filter(tem_oportunidade_aberta=True)
        elif tem_oportunidade_aberta is False:
            queryset = queryset.filter(tem_oportunidade_aberta=False)

        apenas_upgrade = parse_bool_param(self.request.query_params.get('apenas_upgrade'))
        if apenas_upgrade is True:
            queryset = queryset.filter(
                status_cliente=Conta.STATUS_CLIENTE_ATIVO,
                tem_oportunidade_upgrade=True,
            )

        return queryset

    def perform_create(self, serializer):
        user = self.request.user
        canal = serializer.validated_data.get('canal')
        
        # Se não for ADMIN ou não forneceu canal, usa o do usuário
        if user.perfil != 'ADMIN' or not canal:
            canal = user.canal
            
        serializer.save(proprietario=user, canal=canal)
    
    @action(detail=True, methods=['get'])
    def contatos(self, request, pk=None):
        """Lista contatos da conta"""
        conta = self.get_object()
        contatos = conta.contatos.all()
        serializer = ContatoSerializer(contatos, many=True, context={'request': request})
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def oportunidades(self, request, pk=None):
        """Lista oportunidades da conta"""
        conta = self.get_object()
        from django.db.models import Q
        from .models import Oportunidade
        oportunidades = Oportunidade.objects.filter(Q(conta=conta) | Q(empresas=conta)).distinct()
        serializer = OportunidadeSerializer(oportunidades, many=True, context={'request': request})
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def buscar_cnpj(self, request):
        """Busca dados de empresa por CNPJ via ReceitaWS (proxy para evitar CORS)"""
        import requests
        import re
        
        cnpj = request.query_params.get('cnpj', '')
        cnpj_limpo = re.sub(r'\D', '', cnpj)
        
        if len(cnpj_limpo) != 14:
            return Response({'status': 'ERROR', 'message': 'CNPJ deve ter 14 dígitos'}, status=400)
        
        try:
            response = requests.get(
                f'https://www.receitaws.com.br/v1/cnpj/{cnpj_limpo}',
                timeout=30
            )
            data = response.json()
            return Response(data)
        except requests.Timeout:
            return Response({'status': 'ERROR', 'message': 'Tempo limite excedido. Tente novamente.'}, status=504)
        except Exception as e:
            logger.error(f"Erro ao consultar CNPJ {cnpj_limpo}: {str(e)}")
            return Response({'status': 'ERROR', 'message': 'Erro ao consultar CNPJ'}, status=500)

    @action(detail=False, methods=['get'])
    def sem_endereco(self, request):
        """Lista contas com endereço incompleto (qualquer campo vazio)"""
        qs = self.get_queryset().filter(
            Q(endereco='') | Q(endereco__isnull=True) |
            Q(cidade='') | Q(cidade__isnull=True) |
            Q(estado='') | Q(estado__isnull=True) |
            Q(cep='') | Q(cep__isnull=True)
        ).order_by('nome_empresa')
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def atualizar_lote(self, request):
        """Atualiza campos de endereço em lote via lista de {id, ...campos}"""
        CAMPOS_PERMITIDOS = {'nome_empresa', 'marca', 'endereco', 'cidade', 'estado', 'cep', 'telefone_principal', 'email', 'setor'}
        updates = request.data
        if not isinstance(updates, list):
            return Response({'error': 'Esperado uma lista de atualizações'}, status=400)

        qs = self.get_queryset()
        resultados = []
        for item in updates:
            conta_id = item.get('id')
            if not conta_id:
                resultados.append({'id': None, 'status': 'error', 'msg': 'id ausente'})
                continue
            try:
                conta = qs.get(pk=conta_id)
                dados = {k: v for k, v in item.items() if k in CAMPOS_PERMITIDOS and v}
                for campo, valor in dados.items():
                    setattr(conta, campo, valor)
                conta.save(update_fields=list(dados.keys()))
                resultados.append({'id': conta_id, 'status': 'ok'})
            except Exception as e:
                logger.error(f"Erro ao atualizar conta {conta_id} em lote: {e}")
                resultados.append({'id': conta_id, 'status': 'error', 'msg': str(e)})

        return Response(resultados)


class ContaMapaView(APIView):
    """Retorna contas com dados de localização para o mapa do Brasil"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user

        # Permissões de visibilidade
        if user.perfil == 'ADMIN':
            qs = Conta.objects.all()
        else:
            qs = Conta.objects.filter(
                Q(proprietario__canal=user.canal) | Q(canal=user.canal)
            ).distinct()

        # Filtros opcionais
        canal_id = request.query_params.get('canal_id')
        if canal_id:
            qs = qs.filter(canal_id=canal_id)

        status_filter = request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status_cliente=status_filter)

        # Apenas contas com estado informado
        qs = qs.filter(estado__isnull=False).exclude(estado='').select_related('canal')

        data = [
            {
                'id': c.id,
                'nome_empresa': c.nome_empresa,
                'endereco': c.endereco or '',
                'cidade': c.cidade or '',
                'estado': c.estado,
                'canal_id': c.canal_id,
                'canal_nome': c.canal.nome if c.canal else 'Sem Canal',
                'canal_cor': c.canal.cor if c.canal else '#64748b',
                'status_cliente': c.status_cliente,
            }
            for c in qs
        ]
        return Response(data)


class MapaCanalView(APIView):
    """Retorna contas e oportunidades do canal do usuário logado para o mapa"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user

        # Determina o canal do usuário
        if user.perfil == 'ADMIN':
            # Admin sem canal específico → retorna vazio (use ContaMapaView)
            canal = None
        else:
            canal = user.canal

        if not canal:
            return Response({
                'canal': None,
                'contas': [],
                'oportunidades': [],
            })

        # Contas do canal com estado — apenas Clientes Ativos no mapa
        contas_qs = Conta.objects.filter(
            Q(canal=canal) | Q(proprietario__canal=canal)
        ).filter(
            estado__isnull=False,
            status_cliente='CLIENTE_ATIVO'
        ).exclude(estado='').distinct()

        contas = [
            {
                'id': c.id,
                'nome_empresa': c.nome_empresa,
                'endereco': c.endereco or '',
                'cidade': c.cidade or '',
                'estado': c.estado.upper() if c.estado else '',
                'status_cliente': c.status_cliente,
                'canal_id': canal.id,
                'canal_nome': canal.nome,
                'canal_cor': canal.cor,
            }
            for c in contas_qs
        ]

        # Oportunidades do canal — apenas funil de Vendas em estágio Aberto (de Entrada)
        opps_qs = Oportunidade.objects.filter(
            Q(canal=canal) | Q(proprietario__canal=canal)
        ).filter(
            Q(conta__estado__isnull=False) | Q(empresas__estado__isnull=False),
            funil__tipo=Funil.TIPO_VENDAS,
            estagio__tipo=EstagioFunil.TIPO_ABERTO,
        ).select_related('conta', 'estagio').prefetch_related('empresas').distinct()

        opps = []
        for opp in opps_qs:
            # Pega o estado e cidade da conta principal ou da primeira empresa vinculada
            estado = None
            cidade = None
            empresa_nome = None
            
            endereco = ''
            if opp.conta and opp.conta.estado:
                estado = opp.conta.estado.upper()
                cidade = opp.conta.cidade
                endereco = opp.conta.endereco or ''
                empresa_nome = opp.conta.nome_empresa
            else:
                for emp in opp.empresas.filter(estado__isnull=False).exclude(estado='')[:1]:
                    estado = emp.estado.upper()
                    cidade = emp.cidade
                    endereco = emp.endereco or ''
                    empresa_nome = emp.nome_empresa

            if not estado:
                continue

            opps.append({
                'id': opp.id,
                'nome': opp.nome,
                'estado': estado,
                'cidade': cidade or '',
                'endereco': endereco,
                'empresa_nome': empresa_nome or '',
                'valor_estimado': float(opp.valor_estimado) if opp.valor_estimado else 0,
                'estagio_nome': opp.estagio.nome,
                'estagio_cor': opp.estagio.cor,
                'canal_id': canal.id,
                'canal_nome': canal.nome,
                'canal_cor': canal.cor,
            })

        return Response({
            'canal': {
                'id': canal.id,
                'nome': canal.nome,
                'estado': canal.estado,
                'cor': canal.cor,
            },
            'contas': contas,
            'oportunidades': opps,
        })



class TipoContatoViewSet(viewsets.ModelViewSet):
    """ViewSet para Tipos de Contato (CRUD apenas Admin, leitura para autenticados)"""
    queryset = TipoContato.objects.all()
    serializer_class = TipoContatoSerializer
    
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [permissions.IsAuthenticated()]
        return [IsAdminUser()]


class TipoRedeSocialViewSet(viewsets.ModelViewSet):
    """ViewSet para Tipos de Redes Sociais (CRUD apenas Admin, leitura para autenticados)"""
    serializer_class = TipoRedeSocialSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['ordem', 'nome']
    ordering = ['ordem', 'nome']
    
    def get_queryset(self):
        # Retorna todos os tipos ativos (filtro no Python em vez de na query para compatibilidade)
        return TipoRedeSocial.objects.all().order_by('ordem', 'nome')
    
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [permissions.IsAuthenticated()]
        return [IsAdminUser()]


class TagViewSet(viewsets.ModelViewSet):
    """ViewSet para Tags (leitura e criação para autenticados, edição e exclusão apenas Admin)"""
    serializer_class = TagSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['nome']
    ordering_fields = ['nome']

    def get_queryset(self):
        from .models import Tag
        return Tag.objects.all().order_by('nome')

    def get_permissions(self):
        # Qualquer usuário autenticado pode listar, buscar e CRIAR tags
        if self.action in ['list', 'retrieve', 'create']:
            return [permissions.IsAuthenticated()]
        # Apenas admin pode editar ou excluir
        return [IsAdminUser()]


class ContatoViewSet(viewsets.ModelViewSet):
    """ViewSet para Contatos"""
    serializer_class = ContatoSerializer
    
    @action(detail=False, methods=['get'], url_path='checar_telefone')
    def checar_telefone(self, request):
        telefone = request.query_params.get('telefone', '')
        tel_digits = ''.join(filter(str.isdigit, telefone))
        if len(tel_digits) < 8:
            return Response({'exists': False})
            
        from django.db.models import Q
        
        # 1. Busca exata (mesma string enviada pelo front)
        contato = self.get_queryset().filter(Q(telefone=telefone) | Q(celular=telefone)).first()
        
        # 2. Busca parcial (pega registros que têm a mesma região ou dígitos na base e faz filtro Python)
        if not contato:
            qs = self.get_queryset().filter(
                Q(telefone__icontains=tel_digits[-8:]) | 
                Q(celular__icontains=tel_digits[-8:])
            )
            for c in qs:
                t_limpo = ''.join(filter(str.isdigit, c.telefone or ''))
                c_limpo = ''.join(filter(str.isdigit, c.celular or ''))
                if tel_digits[-8:] in t_limpo or tel_digits[-8:] in c_limpo:
                    contato = c
                    break

        # 3. Busca nas tabelas secundárias
        if not contato:
            from .models import ContatoTelefone
            ct_qs = ContatoTelefone.objects.filter(numero=telefone)
            ct = ct_qs.filter(contato__in=self.get_queryset()).first()
            if not ct:
                ct_qs_parcial = ContatoTelefone.objects.filter(numero__icontains=tel_digits[-8:])
                for ctt in ct_qs_parcial.filter(contato__in=self.get_queryset()):
                    n_limpo = ''.join(filter(str.isdigit, ctt.numero or ''))
                    if tel_digits[-8:] in n_limpo:
                        ct = ctt
                        break
            if ct:
                contato = ct.contato

        if contato:
            from django.contrib.contenttypes.models import ContentType
            ct = ContentType.objects.get_for_model(contato)
            return Response({
                'exists': True,
                'contato': {
                    'id': contato.id,
                    'nome': contato.nome,
                    'conta_nome': contato.conta.nome_empresa if contato.conta else None,
                    'content_type_id': ct.id
                }
            })
            
        return Response({'exists': False})
    permission_classes = [HierarchyPermission]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['conta', 'tipo', 'canal', 'tags', 'proprietario']
    search_fields = ['nome', 'email', 'conta__nome_empresa', 'telefone', 'celular', 'telefones__numero']
    ordering_fields = ['nome', 'data_criacao']

    def get_queryset(self):
        user = self.request.user

        if user.perfil == 'ADMIN':
            queryset = Contato.objects.all()
        elif user.perfil in ['RESPONSAVEL', 'VENDEDOR']:
            # Vê contatos criados por ele/vendedores ou vinculados ao seu canal
            queryset = Contato.objects.filter(
                Q(proprietario__canal=user.canal) | Q(canal=user.canal)
            ).distinct()
        else:
            queryset = Contato.objects.filter(proprietario=user)

        queryset = queryset.select_related('proprietario', 'canal', 'conta', 'tipo_contato')

        # Filtro customizado para tipo_contato (suporta string vazia = NULL)
        tipo_contato_param = self.request.query_params.get('tipo_contato', None)
        if tipo_contato_param is not None:
            if tipo_contato_param == '':
                # String vazia = filtrar por contatos sem tipo
                queryset = queryset.filter(tipo_contato__isnull=True)
            else:
                # Filtrar por tipo específico
                queryset = queryset.filter(tipo_contato=tipo_contato_param)

        conta_status_cliente = self.request.query_params.get('conta_status_cliente')
        if conta_status_cliente in {
            Conta.STATUS_PROSPECT,
            Conta.STATUS_CLIENTE_ATIVO,
            Conta.STATUS_INATIVO,
        }:
            queryset = queryset.filter(conta__status_cliente=conta_status_cliente)

        apenas_upgrade = parse_bool_param(self.request.query_params.get('apenas_upgrade'))
        if apenas_upgrade is True:
            opps_abertas_vendas_conta = Oportunidade.objects.filter(
                Q(conta_id=OuterRef('conta_id')) | Q(empresas__id=OuterRef('conta_id')),
                estagio__tipo=EstagioFunil.TIPO_ABERTO,
                funil__tipo=Funil.TIPO_VENDAS,
            )
            queryset = queryset.annotate(
                conta_tem_oportunidade_upgrade=Exists(opps_abertas_vendas_conta)
            ).filter(
                conta__status_cliente=Conta.STATUS_CLIENTE_ATIVO,
                conta_tem_oportunidade_upgrade=True,
            )

        return queryset

    def perform_create(self, serializer):
        user = self.request.user
        canal = serializer.validated_data.get('canal')

        # Se não for ADMIN ou não forneceu canal, usa o do usuário
        if user.perfil != 'ADMIN' or not canal:
            canal = user.canal

        serializer.save(proprietario=user, canal=canal)

    @action(detail=False, methods=['get'])
    def estatisticas(self, request):
        """
        Retorna estatísticas de contatos agrupadas por tipo de contato e por canal
        GET /api/contatos/estatisticas/
        """
        queryset = self.get_queryset()

        # Total geral de contatos
        total = queryset.count()

        # Contatos agrupados por tipo_contato
        stats_por_tipo = queryset.values(
            'tipo_contato__id', 'tipo_contato__nome', 'tipo_contato__emoji'
        ).annotate(
            total=Count('id')
        ).order_by('-total')

        # Formatar resultado de tipos
        tipos = []
        for stat in stats_por_tipo:
            # Garante que sempre há um nome válido
            tipo_id = stat['tipo_contato__id']
            tipo_nome = stat['tipo_contato__nome']
            tipo_emoji = stat['tipo_contato__emoji']

            # Se o nome é None/vazio, é um contato sem tipo
            if not tipo_nome:
                tipo_nome = 'Sem Tipo'
                tipo_id = 'null'  # Usa string 'null' para diferenciar de undefined no frontend
                tipo_emoji = '👤'  # Emoji padrão para contatos sem tipo

            tipos.append({
                'id': tipo_id,
                'nome': tipo_nome,
                'emoji': tipo_emoji or '👤',  # Emoji padrão se não tiver
                'total': stat['total']
            })

        # Contatos agrupados por canal
        stats_por_canal = queryset.values(
            'canal__id', 'canal__nome'
        ).annotate(
            total=Count('id')
        ).order_by('-total')

        # Formatar resultado de canais
        canais = []
        for stat in stats_por_canal:
            canal_id = stat['canal__id']
            canal_nome = stat['canal__nome']

            # Se o nome é None/vazio, é um contato sem canal
            if not canal_nome:
                canal_nome = 'Sem Canal'
                canal_id = 'null'

            canais.append({
                'id': canal_id,
                'nome': canal_nome,
                'total': stat['total']
            })

        return Response({
            'total': total,
            'por_tipo': tipos,
            'por_canal': canais
        })


class EstagioFunilViewSet(viewsets.ModelViewSet):
    """ViewSet para Definições de Estágios (Admin para CRUD, todos podem ler)"""
    queryset = EstagioFunil.objects.all()
    serializer_class = EstagioFunilSerializer
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['tipo', 'funis_vinculados']
    search_fields = ['nome']
    ordering_fields = ['nome']
    
    def get_permissions(self):
        """Admin pode modificar, qualquer usuário autenticado pode ler"""
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsAdminUser()]
        # Estágios são definições globais do sistema, qualquer usuário autenticado pode ler
        return [permissions.IsAuthenticated()]


class OportunidadeViewSet(viewsets.ModelViewSet):
    """ViewSet para Oportunidades"""
    serializer_class = OportunidadeSerializer
    permission_classes = [HierarchyPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = {
        'estagio': ['exact'],
        'estagio__tipo': ['exact', 'in'],
        'conta': ['exact'],
        'canal': ['exact'],
        'funil': ['exact'],
        'funil__tipo': ['exact'],
        'contatos': ['exact'],
        'empresas': ['exact'],
        'origem': ['exact'],
        'tags': ['exact'],
    }
    search_fields = ['nome', 'conta__nome_empresa', 'empresas__nome_empresa', 'contatos__nome']
    ordering_fields = ['nome', 'conta__nome_empresa', 'valor_estimado', 'data_fechamento_esperada', 'data_criacao']
    
    def get_queryset(self):
        """Aplica filtros de hierarquia e funis de acesso"""
        user = self.request.user
        
        if user.perfil == 'ADMIN':
            queryset = Oportunidade.objects.all()
        else:
            # Pega todos os funis aos quais o usuário tem acesso (direto, via canal ou globais)
            funis_visiveis = Funil.objects.filter(
                Q(usuarios=user) | 
                Q(usuarios__canal=user.canal) |
                Q(usuarios__isnull=True)
            ).distinct()
            
            if user.perfil == 'RESPONSAVEL':
                # Vê do seu canal (vínculo direto ou via proprietário) AND funis visíveis
                queryset = Oportunidade.objects.filter(
                    Q(canal=user.canal) | Q(proprietario__canal=user.canal)
                ).filter(funil__in=funis_visiveis).distinct()
            else: # VENDEDOR
                if user.canal:
                    queryset = Oportunidade.objects.filter(
                        Q(canal=user.canal) | Q(proprietario__canal=user.canal),
                        funil__in=funis_visiveis
                    ).distinct()
                else:
                    # Fallback: se não tem canal, vê só as próprias
                    queryset = Oportunidade.objects.filter(
                        proprietario=user,
                        funil__in=funis_visiveis
                    )
            
        return queryset.select_related(
            'funil', 'estagio', 'conta', 'contato_principal', 'proprietario'
        ).prefetch_related(
            'oportunidadeadicional_set', 'contatos', 'empresas', 'anexos'
        )

    def perform_create(self, serializer):
        # Se não forneceu estágio, busca o padrão do funil
        estagio = serializer.validated_data.get('estagio')
        funil = serializer.validated_data.get('funil')
        
        if not estagio and funil:
            vinculo_padrao = FunilEstagio.objects.filter(funil=funil, is_padrao=True).first()
            if not vinculo_padrao:
                # Se não houver marcado como padrão, pega o de menor ordem
                vinculo_padrao = FunilEstagio.objects.filter(funil=funil).order_by('ordem').first()
            
            if vinculo_padrao:
                estagio = vinculo_padrao.estagio
        
        # Atribui canal automaticamente se o usuário tiver um
        canal = serializer.validated_data.get('canal')
        if not canal and self.request.user.canal:
            canal = self.request.user.canal
            
        # Garante que a conta da oportunidade também esteja vinculada ao mesmo canal
        # apenas se a conta ainda não tiver um canal definido
        conta = serializer.validated_data.get('conta')
        if conta and canal and not conta.canal:
            conta.canal = canal
            conta.save()
            
        serializer.save(
            proprietario=self.request.user, 
            estagio=estagio,
            canal=canal
        )
        
        # Registra histórico de criação (primeiro estágio)
        from .models import HistoricoEstagio, FunilEstagio as FE
        oportunidade = serializer.instance
        if oportunidade.estagio:
            fe_novo = FE.objects.filter(
                funil=oportunidade.funil,
                estagio=oportunidade.estagio
            ).first()
            
            HistoricoEstagio.objects.create(
                tipo_objeto=HistoricoEstagio.TIPO_OPORTUNIDADE,
                oportunidade=oportunidade,
                estagio_anterior=None,
                estagio_novo=fe_novo,
                nome_estagio_anterior=None,
                nome_estagio_novo=oportunidade.estagio.nome,
                usuario=self.request.user,
                observacao='Criação da oportunidade'
            )
    
    def perform_update(self, serializer):
        """Registra histórico quando o estágio é alterado"""
        from .models import HistoricoEstagio, FunilEstagio
        
        instance = self.get_object()
        estagio_anterior = instance.estagio
        
        # Salva as mudanças
        serializer.save()
        
        # Verifica se o estágio mudou
        estagio_novo = serializer.instance.estagio
        if estagio_anterior != estagio_novo:
            # Busca o FunilEstagio correspondente para usar no histórico
            fe_anterior = FunilEstagio.objects.filter(
                funil=instance.funil,
                estagio=estagio_anterior
            ).first() if estagio_anterior else None
            
            fe_novo = FunilEstagio.objects.filter(
                funil=serializer.instance.funil,
                estagio=estagio_novo
            ).first() if estagio_novo else None
            
            HistoricoEstagio.objects.create(
                tipo_objeto=HistoricoEstagio.TIPO_OPORTUNIDADE,
                oportunidade=serializer.instance,
                estagio_anterior=fe_anterior,
                estagio_novo=fe_novo,
                nome_estagio_anterior=estagio_anterior.nome if estagio_anterior else None,
                nome_estagio_novo=estagio_novo.nome if estagio_novo else None,
                usuario=self.request.user
            )
    
    @staticmethod
    def _normalize_header(value):
        text = str(value or '').strip().lower()
        text = ''.join(
            ch for ch in unicodedata.normalize('NFD', text)
            if unicodedata.category(ch) != 'Mn'
        )
        return re.sub(r'\s+', ' ', text)

    @staticmethod
    def _normalize_digits(value):
        return re.sub(r'\D', '', str(value or ''))

    @staticmethod
    def _normalize_phone_from_import(value):
        """Normaliza telefone vindo da planilha para apenas dígitos, inclusive notação científica."""
        if value is None:
            return ''

        if isinstance(value, int):
            return str(value)

        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return re.sub(r'\D', '', format(value, 'f'))

        text = str(value).strip()
        if not text:
            return ''

        if re.match(r'^[+-]?\d+(\.\d+)?[eE][+-]?\d+$', text):
            try:
                decimal_value = Decimal(text)
                plain = format(decimal_value, 'f')
                integer_part = plain.split('.')[0]
                return re.sub(r'\D', '', integer_part)
            except (InvalidOperation, ValueError):
                pass

        return re.sub(r'\D', '', text)

    @staticmethod
    def _normalize_import_text(value):
        text = str(value or '').strip()
        if not text:
            return ''

        text = text.replace('\u00a0', ' ')
        if 'Ã' in text or 'Â' in text or '�' in text:
            try:
                repaired = text.encode('latin1').decode('utf-8')
                if repaired.strip():
                    text = repaired.strip()
            except (UnicodeEncodeError, UnicodeDecodeError):
                pass

        return text

    @classmethod
    def _get_row_value(cls, row_values, headers, aliases, fixed_index=None):
        for alias in aliases:
            idx = headers.get(cls._normalize_header(alias))
            if idx is not None and idx < len(row_values):
                val = row_values[idx]
                if val is not None and str(val).strip() != '':
                    return cls._normalize_import_text(val)
        if fixed_index is not None and fixed_index < len(row_values):
            val = row_values[fixed_index]
            if val is not None and str(val).strip() != '':
                return cls._normalize_import_text(val)
        return ''

    def _resolve_canal_import(self, request):
        user = request.user
        canal_id = request.data.get('canal_id')

        if user.perfil == 'ADMIN':
            if canal_id:
                canal = Canal.objects.filter(id=canal_id).first()
                if not canal:
                    return None, Response({'error': 'Canal informado não encontrado.'}, status=400)
                return canal, None
            if user.canal:
                return user.canal, None
            return None, Response({'error': 'Informe canal_id para importação.'}, status=400)

        if user.canal:
            return user.canal, None

        return None, Response({'error': 'Usuário sem canal vinculado para importar.'}, status=403)

    def _resolve_funil_estagio(self, canal):
        funil = canal.funil_padrao if canal else None
        estagio = canal.estagio_inicial if canal else None

        if not funil:
            funil = Funil.objects.filter(tipo=Funil.TIPO_VENDAS, is_active=True).first()

        if not funil:
            return None, None

        if not estagio:
            vinculo = FunilEstagio.objects.filter(funil=funil, is_padrao=True).first() or \
                      FunilEstagio.objects.filter(funil=funil).order_by('ordem').first()
            estagio = vinculo.estagio if vinculo else None

        return funil, estagio

    def _resolve_funil_estagio_por_tipo(self, funil_tipo):
        funil = Funil.objects.filter(tipo=funil_tipo, is_active=True).first()
        if not funil:
            return None, None

        vinculo = FunilEstagio.objects.filter(funil=funil, is_padrao=True).first() or \
                  FunilEstagio.objects.filter(funil=funil).order_by('ordem').first()
        estagio = vinculo.estagio if vinculo else None
        return funil, estagio

    @staticmethod
    def _get_request_ip(request):
        forwarded = request.META.get('HTTP_X_FORWARDED_FOR', '')
        if forwarded:
            return forwarded.split(',')[0].strip()
        return request.META.get('REMOTE_ADDR')

    def _can_access_import_log(self, user, import_log):
        if user.perfil == 'ADMIN':
            return True

        payload = import_log.alteracoes or {}
        canal_payload = payload.get('canal') or {}
        canal_id = canal_payload.get('id')

        return import_log.usuario_id == user.id or (user.canal_id and canal_id == user.canal_id)

    def _build_import_log_response_item(self, import_log):
        payload = import_log.alteracoes or {}
        created_ids = payload.get('created_ids') or {}
        rollback = payload.get('rollback') or {}

        return {
            'lote_id': import_log.id,
            'arquivo': payload.get('arquivo'),
            'canal': payload.get('canal'),
            'funil': payload.get('funil'),
            'estagio': payload.get('estagio'),
            'summary': payload.get('summary') or {},
            'created_counts': {
                'oportunidades': len(created_ids.get('oportunidades_ids') or []),
                'contas': len(created_ids.get('contas_ids') or []),
                'contatos': len(created_ids.get('contatos_ids') or []),
                'origens': len(created_ids.get('origens_ids') or []),
            },
            'revertido': bool(rollback.get('revertido_em')),
            'rollback': rollback,
            'usuario': {
                'id': import_log.usuario_id,
                'nome': import_log.usuario.get_full_name() if import_log.usuario else None,
            },
            'criado_em': import_log.timestamp,
        }

    @action(detail=False, methods=['post'])
    def importar_xlsx(self, request):
        """Importa oportunidades via planilha XLSX (Conta + Contato + Oportunidade)."""
        try:
            from openpyxl import load_workbook
        except Exception as e:
            return Response(
                {'error': f'Falha ao importar openpyxl: {e}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        file = request.FILES.get('file')
        dry_run = str(request.data.get('dry_run', 'true')).lower() in ['1', 'true', 'yes', 'sim']

        if not file:
            return Response({'error': 'Arquivo é obrigatório.'}, status=400)

        if not file.name.lower().endswith('.xlsx'):
            return Response({'error': 'Formato inválido. Envie um arquivo .xlsx'}, status=400)

        canal, canal_error = self._resolve_canal_import(request)
        if canal_error:
            return canal_error

        funil, estagio = self._resolve_funil_estagio(canal)
        if not funil:
            return Response({'error': 'Nenhum funil de vendas configurado.'}, status=400)
        if not estagio:
            return Response({'error': 'Nenhum estágio inicial encontrado para o funil selecionado.'}, status=400)

        try:
            wb = load_workbook(file, data_only=True)
            ws = wb.active
        except Exception as e:
            return Response({'error': f'Falha ao ler planilha: {e}'}, status=400)

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return Response({'error': 'Planilha sem cabeçalho.'}, status=400)

        headers = {
            self._normalize_header(col): idx
            for idx, col in enumerate(header_row)
            if col is not None and str(col).strip()
        }

        summary = {
            'processadas': 0,
            'criadas': 0,
            'atualizadas': 0,
            'puladas': 0,
            'erros': 0,
            'contas_criadas': 0,
            'contatos_criados': 0,
            'origens_criadas': 0,
        }
        rows_result = []
        created_ids = {
            'oportunidades_ids': set(),
            'contas_ids': set(),
            'contatos_ids': set(),
            'origens_ids': set(),
        }

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_values = list(row or [])

            nome = self._get_row_value(row_values, headers, ['Nome', 'Contato', 'Nome Contato'], fixed_index=7)
            cnpj_raw = self._get_row_value(row_values, headers, ['CNPJ'], fixed_index=1)
            razao_social = self._get_row_value(row_values, headers, ['Razao Social', 'Razão Social'], fixed_index=2)
            nome_fantasia = self._get_row_value(row_values, headers, ['Nome Fantasia', 'Fantasia'], fixed_index=3)
            segmento = self._get_row_value(row_values, headers, ['Segmento'], fixed_index=4)
            cidade = self._get_row_value(row_values, headers, ['Nome Cidade', 'Cidade'], fixed_index=5)
            estado = self._get_row_value(row_values, headers, ['UF', 'Uf', 'Estado'], fixed_index=6)
            email = self._get_row_value(row_values, headers, ['Email', 'E-mail'], fixed_index=8)
            telefone = self._get_row_value(row_values, headers, ['Telefone', 'Celular', 'WhatsApp'], fixed_index=9)
            fonte_nome = self._get_row_value(row_values, headers, ['Fonte', 'Origem'], fixed_index=10)
            estande = self._get_row_value(row_values, headers, ['Estande'], fixed_index=0)

            if not any([nome, cnpj_raw, razao_social, nome_fantasia, email, telefone, fonte_nome]):
                continue

            summary['processadas'] += 1

            if not nome:
                summary['erros'] += 1
                rows_result.append({'linha': row_idx, 'status': 'erro', 'mensagem': 'Nome do contato obrigatório.'})
                continue

            empresa_nome = (nome_fantasia or razao_social).strip()
            if not empresa_nome and not cnpj_raw:
                summary['erros'] += 1
                rows_result.append({'linha': row_idx, 'status': 'erro', 'mensagem': 'Empresa sem CNPJ e sem nome.'})
                continue

            cnpj = self._normalize_digits(cnpj_raw)
            if cnpj and len(cnpj) != 14:
                cnpj = ''
            telefone = self._normalize_phone_from_import(telefone)

            try:
                with transaction.atomic():
                    conta = None
                    if cnpj:
                        conta = Conta.objects.filter(cnpj=cnpj).first()

                    if not conta and empresa_nome:
                        conta = Conta.objects.filter(nome_empresa__iexact=empresa_nome).first()

                    if not conta and not dry_run:
                        conta = Conta.objects.create(
                            nome_empresa=empresa_nome or f'Empresa {nome}',
                            cnpj=cnpj or None,
                            telefone_principal=telefone or None,
                            setor=segmento or None,
                            cidade=cidade or None,
                            estado=(estado or '').upper()[:2] or None,
                            canal=canal,
                            proprietario=request.user
                        )
                        created_ids['contas_ids'].add(conta.id)
                        summary['contas_criadas'] += 1

                    if conta and not dry_run:
                        dirty = False
                        if telefone and not conta.telefone_principal:
                            conta.telefone_principal = telefone
                            dirty = True
                        if segmento and not conta.setor:
                            conta.setor = segmento
                            dirty = True
                        if cidade and not conta.cidade:
                            conta.cidade = cidade
                            dirty = True
                        if estado and not conta.estado:
                            conta.estado = estado.upper()[:2]
                            dirty = True
                        if canal and not conta.canal:
                            conta.canal = canal
                            dirty = True
                        if dirty:
                            conta.save()
                            summary['atualizadas'] += 1

                    contato = None
                    if email:
                        contato = Contato.objects.filter(email__iexact=email).first()

                    if not contato and conta:
                        contato = Contato.objects.filter(nome__iexact=nome, conta=conta).first()

                    if not contato and not dry_run:
                        contato = Contato.objects.create(
                            nome=nome,
                            email=email or None,
                            telefone=telefone or None,
                            celular=telefone or None,
                            conta=conta,
                            canal=canal,
                            proprietario=request.user
                        )
                        created_ids['contatos_ids'].add(contato.id)
                        summary['contatos_criados'] += 1

                    if contato and not dry_run:
                        contato_dirty = False
                        if telefone and not contato.telefone:
                            contato.telefone = telefone
                            contato_dirty = True
                        if telefone and not contato.celular:
                            contato.celular = telefone
                            contato_dirty = True
                        if conta and not contato.conta:
                            contato.conta = conta
                            contato_dirty = True
                        if canal and not contato.canal:
                            contato.canal = canal
                            contato_dirty = True
                        if contato_dirty:
                            contato.save()
                            summary['atualizadas'] += 1

                    origem = None
                    if fonte_nome and not dry_run:
                        origem, created_origem = Origem.objects.get_or_create(
                            nome=fonte_nome.strip(),
                            defaults={'ativo': True}
                        )
                        if created_origem:
                            created_ids['origens_ids'].add(origem.id)
                            summary['origens_criadas'] += 1

                    opp_suffix = (nome_fantasia or razao_social).strip()
                    opp_name = f"{nome} - {opp_suffix}" if opp_suffix else nome
                    opp_name = opp_name[:255]

                    if dry_run:
                        summary['criadas'] += 1
                        rows_result.append({
                            'linha': row_idx,
                            'status': 'preview',
                            'mensagem': 'Linha válida para importação.',
                            'oportunidade_nome': opp_name,
                            'empresa': empresa_nome,
                            'contato': nome,
                            'fonte': fonte_nome,
                        })
                        continue

                    oportunidade = Oportunidade.objects.create(
                        nome=opp_name,
                        conta=conta,
                        contato_principal=contato,
                        funil=funil,
                        estagio=estagio,
                        canal=canal,
                        proprietario=request.user,
                        fonte=fonte_nome or None,
                        origem=origem,
                        descricao=f"Estande: {estande}" if estande else None,
                    )
                    created_ids['oportunidades_ids'].add(oportunidade.id)

                    if contato:
                        oportunidade.contatos.add(contato)
                    if conta:
                        oportunidade.empresas.add(conta)

                    summary['criadas'] += 1
                    rows_result.append({
                        'linha': row_idx,
                        'status': 'ok',
                        'mensagem': 'Importado com sucesso.',
                        'oportunidade_id': oportunidade.id,
                        'oportunidade_nome': oportunidade.nome,
                    })

            except Exception as e:
                summary['erros'] += 1
                rows_result.append({'linha': row_idx, 'status': 'erro', 'mensagem': str(e)})

        if summary['processadas'] == 0:
            return Response({'error': 'Nenhuma linha válida encontrada na planilha.'}, status=400)

        import_lote = None
        if not dry_run:
            created_ids_json = {
                key: sorted(list(value))
                for key, value in created_ids.items()
            }
            import_lote = Log.objects.create(
                usuario=request.user,
                acao=Log.ACAO_IMPORT,
                modelo='IMPORTACAO_XLSX_OPORTUNIDADES',
                objeto_repr=f"Importação XLSX - {file.name}",
                alteracoes={
                    'arquivo': file.name,
                    'canal': {'id': canal.id, 'nome': canal.nome} if canal else None,
                    'funil': {'id': funil.id, 'nome': funil.nome} if funil else None,
                    'estagio': {'id': estagio.id, 'nome': estagio.nome} if estagio else None,
                    'summary': summary,
                    'created_ids': created_ids_json,
                },
                ip_address=self._get_request_ip(request),
                user_agent=(request.META.get('HTTP_USER_AGENT') or '')[:1000] or None,
                observacao='Lote de importação de oportunidades via XLSX',
            )

        return Response({
            'dry_run': dry_run,
            'canal': {'id': canal.id, 'nome': canal.nome} if canal else None,
            'funil': {'id': funil.id, 'nome': funil.nome} if funil else None,
            'estagio': {'id': estagio.id, 'nome': estagio.nome} if estagio else None,
            'summary': summary,
            'rows': rows_result[:200],
            'lote_importacao': {'id': import_lote.id, 'criado_em': import_lote.timestamp} if import_lote else None,
        })

    @action(detail=False, methods=['get'], url_path='importacao_lotes')
    def importacao_lotes(self, request):
        """Lista lotes de importação XLSX de oportunidades."""
        logs = Log.objects.filter(
            acao=Log.ACAO_IMPORT,
            modelo='IMPORTACAO_XLSX_OPORTUNIDADES'
        ).select_related('usuario').order_by('-timestamp')[:100]

        lotes = []
        for import_log in logs:
            if self._can_access_import_log(request.user, import_log):
                lotes.append(self._build_import_log_response_item(import_log))

        return Response({'results': lotes})

    @action(detail=False, methods=['post'], url_path='reverter_importacao')
    def reverter_importacao(self, request):
        """Reverte um lote de importação removendo registros criados no lote, quando seguro."""
        lote_id = request.data.get('lote_id')
        if not lote_id:
            return Response({'error': 'lote_id é obrigatório.'}, status=400)

        try:
            lote_id = int(lote_id)
        except (TypeError, ValueError):
            return Response({'error': 'lote_id inválido.'}, status=400)

        import_log = Log.objects.filter(
            id=lote_id,
            acao=Log.ACAO_IMPORT,
            modelo='IMPORTACAO_XLSX_OPORTUNIDADES'
        ).select_related('usuario').first()

        if not import_log:
            return Response({'error': 'Lote de importação não encontrado.'}, status=404)

        if not self._can_access_import_log(request.user, import_log):
            return Response({'error': 'Sem permissão para reverter este lote.'}, status=403)

        payload = import_log.alteracoes or {}
        rollback_data = payload.get('rollback') or {}
        if rollback_data.get('revertido_em'):
            return Response({'error': 'Este lote já foi revertido.'}, status=400)

        created_ids = payload.get('created_ids') or {}
        oportunidades_ids = created_ids.get('oportunidades_ids') or []
        contatos_ids = created_ids.get('contatos_ids') or []
        contas_ids = created_ids.get('contas_ids') or []
        origens_ids = created_ids.get('origens_ids') or []

        rollback_summary = {
            'oportunidades_excluidas': 0,
            'contatos_excluidos': 0,
            'contatos_preservados': 0,
            'contas_excluidas': 0,
            'contas_preservadas': 0,
            'origens_excluidas': 0,
            'origens_preservadas': 0,
        }

        with transaction.atomic():
            if oportunidades_ids:
                qs_oportunidades = Oportunidade.objects.filter(id__in=oportunidades_ids)
                rollback_summary['oportunidades_excluidas'] = qs_oportunidades.count()
                qs_oportunidades.delete()

            for contato in Contato.objects.filter(id__in=contatos_ids):
                if contato.oportunidades.exists() or contato.oportunidades_principais_contato.exists():
                    rollback_summary['contatos_preservados'] += 1
                    continue
                contato.delete()
                rollback_summary['contatos_excluidos'] += 1

            for conta in Conta.objects.filter(id__in=contas_ids):
                has_rel = (
                    conta.oportunidades.exists() or
                    conta.oportunidades_diretas.exists() or
                    conta.contatos.exists()
                )
                if has_rel:
                    rollback_summary['contas_preservadas'] += 1
                    continue
                conta.delete()
                rollback_summary['contas_excluidas'] += 1

            for origem in Origem.objects.filter(id__in=origens_ids):
                if origem.oportunidades.exists():
                    rollback_summary['origens_preservadas'] += 1
                    continue
                origem.delete()
                rollback_summary['origens_excluidas'] += 1

            payload['rollback'] = {
                'revertido_em': timezone.now().isoformat(),
                'revertido_por': {
                    'id': request.user.id,
                    'nome': request.user.get_full_name() or request.user.username,
                },
                'summary': rollback_summary,
            }
            import_log.alteracoes = payload
            import_log.observacao = 'Lote de importação revertido.'
            import_log.save(update_fields=['alteracoes', 'observacao'])

            Log.objects.create(
                usuario=request.user,
                acao=Log.ACAO_DELETE,
                modelo='IMPORTACAO_XLSX_OPORTUNIDADES',
                objeto_id=import_log.id,
                objeto_repr=import_log.objeto_repr,
                alteracoes={'lote_id': import_log.id, 'rollback_summary': rollback_summary},
                ip_address=self._get_request_ip(request),
                user_agent=(request.META.get('HTTP_USER_AGENT') or '')[:1000] or None,
                observacao='Reversão manual de lote de importação XLSX',
            )

        return Response({
            'success': True,
            'lote_id': import_log.id,
            'rollback_summary': rollback_summary,
        })

    @action(detail=False, methods=['get'])
    def kanban(self, request):
        """Retorna oportunidades agrupadas por estágio para visão Kanban"""
        funil_id = request.query_params.get('funil_id')
        estagio_tipo = request.query_params.get('estagio_tipo') # Novo filtro de status
        
        queryset = self.get_queryset()
        
        # Filtro de status (opcional, se não informado mostra todos)
        if estagio_tipo:
            queryset = queryset.filter(estagio__tipo=estagio_tipo)
            
        if funil_id:
            queryset = queryset.filter(funil_id=funil_id)
        else:
            # Se não informou, tenta o primeiro funil de oportunidades do usuário
            user_funis = request.user.funis_acesso.filter(tipo=Funil.TIPO_VENDAS) if request.user.perfil != 'ADMIN' else Funil.objects.filter(tipo=Funil.TIPO_VENDAS)
            funil = user_funis.first()
            if funil:
                queryset = queryset.filter(funil=funil)
                
        queryset = queryset.select_related('conta', 'contato_principal', 'estagio', 'proprietario')
        serializer = OportunidadeKanbanSerializer(queryset, many=True)
        
        # Agrupa por estágios do funil selecionado via tabela de ligação (todos os tipos)
        funil_selecionado_id = funil_id or (funil.id if 'funil' in locals() and funil else None)
        if funil_selecionado_id:
            vinculos = FunilEstagio.objects.filter(funil_id=funil_selecionado_id).select_related('estagio').order_by('ordem')
        else:
            # Fallback se não houver funil (não deveria acontecer no Kanban novo)
            vinculos = FunilEstagio.objects.all().select_related('estagio').order_by('funil', 'ordem')
        
        kanban_data = []
        for vinculo in vinculos:
            oportunidades = [
                opp for opp in serializer.data
                if int(opp.get('estagio_id') or opp.get('estagio') or 0) == vinculo.estagio_id
            ]
            # Montamos o objeto de estágio como o frontend espera
            estagio_data = EstagioFunilSerializer(vinculo.estagio).data
            estagio_data['ordem'] = vinculo.ordem
            estagio_data['is_padrao'] = vinculo.is_padrao

            kanban_data.append({
                'estagio': estagio_data,
                'oportunidades': oportunidades
            })
        
        return Response(kanban_data)
    
    @action(detail=True, methods=['patch'])
    def mudar_estagio(self, request, pk=None):
        """Muda o estágio de uma oportunidade (usado no drag-and-drop do Kanban)"""
        from .models import HistoricoEstagio
        
        oportunidade = self.get_object()
        estagio_anterior = oportunidade.estagio
        novo_estagio_id = request.data.get('estagio_id')
        
        if not novo_estagio_id:
            return Response(
                {'error': 'estagio_id é obrigatório'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            with transaction.atomic():
                novo_estagio = EstagioFunil.objects.get(id=novo_estagio_id)
                oportunidade.estagio = novo_estagio
                
                # Se mudou para estágio fechado, registra a data
                if novo_estagio.tipo in [EstagioFunil.TIPO_GANHO, EstagioFunil.TIPO_PERDIDO]:
                    oportunidade.data_fechamento_real = timezone.now().date()
                
                oportunidade.save()
                
                # Registra histórico de mudança de estágio
                if estagio_anterior != novo_estagio:
                    # Busca o FunilEstagio correspondente para obter o nome
                    from .models import FunilEstagio
                    fe_anterior = FunilEstagio.objects.filter(
                        funil=oportunidade.funil,
                        estagio=estagio_anterior
                    ).select_related('estagio').first() if estagio_anterior else None
                    
                    fe_novo = FunilEstagio.objects.filter(
                        funil=oportunidade.funil,
                        estagio=novo_estagio
                    ).select_related('estagio').first()
                    
                    HistoricoEstagio.objects.create(
                        tipo_objeto=HistoricoEstagio.TIPO_OPORTUNIDADE,
                        oportunidade=oportunidade,
                        estagio_anterior=fe_anterior,
                        estagio_novo=fe_novo,
                        nome_estagio_anterior=fe_anterior.estagio.nome if fe_anterior else None,
                        nome_estagio_novo=fe_novo.estagio.nome if fe_novo else novo_estagio.nome,
                        usuario=request.user
                    )
            
            # Usamos o KanbanSerializer para a resposta, pois é mais leve e o que o Kanban espera
            serializer = OportunidadeKanbanSerializer(oportunidade)
            return Response(serializer.data)
        
        except EstagioFunil.DoesNotExist:
            return Response(
                {'error': 'Estágio não encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            # Captura erros inesperados para evitar 500 silencioso
            return Response(
                {'error': f'Erro ao mudar estágio: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['post'], url_path='converter_em_cliente')
    def converter_em_cliente(self, request, pk=None):
        """Converte a conta da oportunidade para Cliente Ativo/Inativo e move para Pós-Venda."""
        oportunidade = self.get_object()

        if not oportunidade.conta:
            return Response({'error': 'A oportunidade não possui empresa vinculada.'}, status=400)

        status_destino = str(request.data.get('status_cliente') or Conta.STATUS_CLIENTE_ATIVO).upper()
        if status_destino not in [Conta.STATUS_CLIENTE_ATIVO, Conta.STATUS_INATIVO]:
            return Response({'error': 'status_cliente inválido. Use CLIENTE_ATIVO ou INATIVO.'}, status=400)

        funil_pos, estagio_pos = self._resolve_funil_estagio_por_tipo(Funil.TIPO_POS_VENDA)
        if not funil_pos:
            return Response({'error': 'Nenhum funil de Pós-Venda configurado.'}, status=400)
        if not estagio_pos:
            return Response({'error': 'Nenhum estágio inicial encontrado para o funil de Pós-Venda.'}, status=400)

        estagio_destino = estagio_pos
        if status_destino == Conta.STATUS_INATIVO:
            vinculo_inativo = FunilEstagio.objects.filter(
                funil=funil_pos,
                estagio__nome__icontains='inativ'
            ).select_related('estagio').order_by('ordem').first()
            if not vinculo_inativo:
                vinculo_inativo = FunilEstagio.objects.filter(
                    funil=funil_pos,
                    estagio__tipo=EstagioFunil.TIPO_PERDIDO
                ).select_related('estagio').order_by('ordem').first()
            if vinculo_inativo:
                estagio_destino = vinculo_inativo.estagio

        conta = oportunidade.conta
        before_status = conta.status_cliente
        before_ativacao = conta.data_ativacao_cliente
        before_funil_id = oportunidade.funil_id
        before_funil_nome = oportunidade.funil.nome if oportunidade.funil else None
        before_estagio_id = oportunidade.estagio_id
        before_estagio_nome = oportunidade.estagio.nome if oportunidade.estagio else None

        ja_no_destino = oportunidade.funil_id == funil_pos.id and oportunidade.estagio_id == estagio_destino.id

        if conta.status_cliente == status_destino and ja_no_destino:
            return Response({
                'success': True,
                'idempotente': True,
                'conta': {
                    'id': conta.id,
                    'nome_empresa': conta.nome_empresa,
                    'status_cliente': conta.status_cliente,
                    'data_ativacao_cliente': conta.data_ativacao_cliente,
                },
                'oportunidade': {
                    'id': oportunidade.id,
                    'funil_id': oportunidade.funil_id,
                    'funil_nome': oportunidade.funil.nome if oportunidade.funil else None,
                    'estagio_id': oportunidade.estagio_id,
                    'estagio_nome': oportunidade.estagio.nome if oportunidade.estagio else None,
                },
                'mensagem': 'Conta já estava com o status informado. Nenhuma alteração aplicada.',
            })

        with transaction.atomic():
            conta.status_cliente = status_destino
            if status_destino == Conta.STATUS_CLIENTE_ATIVO and not conta.data_ativacao_cliente:
                conta.data_ativacao_cliente = timezone.now()

            conta.save(update_fields=['status_cliente', 'data_ativacao_cliente', 'data_atualizacao'])

            oportunidade.funil = funil_pos
            oportunidade.estagio = estagio_destino
            oportunidade.save(update_fields=['funil', 'estagio', 'data_atualizacao'])

            Log.objects.create(
                usuario=request.user,
                acao=Log.ACAO_UPDATE,
                modelo='Conta',
                objeto_id=conta.id,
                objeto_repr=str(conta),
                alteracoes={
                    'status_cliente': {'antes': before_status, 'depois': conta.status_cliente},
                    'data_ativacao_cliente': {
                        'antes': before_ativacao.isoformat() if before_ativacao else None,
                        'depois': conta.data_ativacao_cliente.isoformat() if conta.data_ativacao_cliente else None,
                    },
                    'origem': {
                        'tipo': 'oportunidade',
                        'oportunidade_id': oportunidade.id,
                        'oportunidade_nome': oportunidade.nome,
                    },
                    'oportunidade_funil': {
                        'antes': {'id': before_funil_id, 'nome': before_funil_nome},
                        'depois': {'id': oportunidade.funil_id, 'nome': oportunidade.funil.nome if oportunidade.funil else None},
                    },
                    'oportunidade_estagio': {
                        'antes': {'id': before_estagio_id, 'nome': before_estagio_nome},
                        'depois': {'id': oportunidade.estagio_id, 'nome': oportunidade.estagio.nome if oportunidade.estagio else None},
                    },
                },
                ip_address=self._get_request_ip(request),
                user_agent=(request.META.get('HTTP_USER_AGENT') or '')[:1000] or None,
                observacao='Conversão de oportunidade em cliente',
            )

        return Response({
            'success': True,
            'idempotente': False,
            'conta': {
                'id': conta.id,
                'nome_empresa': conta.nome_empresa,
                'status_cliente': conta.status_cliente,
                'data_ativacao_cliente': conta.data_ativacao_cliente,
            },
            'oportunidade': {
                'id': oportunidade.id,
                'funil_id': oportunidade.funil_id,
                'funil_nome': oportunidade.funil.nome if oportunidade.funil else None,
                'estagio_id': oportunidade.estagio_id,
                'estagio_nome': oportunidade.estagio.nome if oportunidade.estagio else None,
            },
            'mensagem': 'Conta convertida com sucesso e oportunidade movida para Pós-Venda.',
        })

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Retorna indicadores resumidos das oportunidades (KPIs)"""
        queryset = self.get_queryset()
        
        # Filtros básicos de busca se enviados
        search = request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(nome__icontains=search) | 
                Q(conta__nome_empresa__icontains=search)
            )

        # Filtro por tipo de funil (ex: só VENDAS)
        funil_tipo = request.query_params.get('funil__tipo')
        if funil_tipo:
            queryset = queryset.filter(funil__tipo=funil_tipo)
        
        # Filtro por funil específico
        funil_id = request.query_params.get('funil')
        if funil_id:
            queryset = queryset.filter(funil_id=funil_id)

        data = queryset.aggregate(
            total_valor=Sum('valor_estimado'),
            total_contagem=Count('id'),
            valor_medio=Avg('valor_estimado'),
            valor_ganho=Sum('valor_estimado', filter=Q(estagio__tipo='GANHO')),
            valor_aberto=Sum('valor_estimado', filter=Q(estagio__tipo='ABERTO')),
            contagem_aberta=Count('id', filter=Q(estagio__tipo='ABERTO')),
            contagem_ganha=Count('id', filter=Q(estagio__tipo='GANHO'))
        )
        
        # Formatar valores nulos para 0
        for key in data:
            if data[key] is None:
                data[key] = 0
        
        return Response(data)

    @action(detail=True, methods=['get'])
    def gerar_texto_faturamento(self, request, pk=None):
        """Gera o texto formatado para faturamento"""
        opp = self.get_object()
        
        if not opp.plano:
            return Response({'error': 'Nenhum plano selecionado para esta oportunidade'}, status=status.HTTP_400_BAD_REQUEST)
            
        adicionais_texto = ""
        adicionais = opp.oportunidadeadicional_set.all()
        if adicionais.exists():
            adicionais_texto = "\n\nRecursos Adicionais Incluídos:\n"
            for ad in adicionais:
                adicionais_texto += f"• {ad.quantidade}x {ad.adicional.nome}\n"

        params = {
            'plano_nome': opp.plano.nome,
            'periodo': opp.get_periodo_pagamento_display(),
            'recursos': "\n".join(opp.plano.recursos),
            'adicionais': adicionais_texto,
            'cortesia': opp.cortesia or "Nenhuma cortesia registrada",
            'cliente_nome': opp.contato_principal.nome if opp.contato_principal else opp.conta.nome_empresa,
            'whatsapp': opp.contato_principal.telefone if opp.contato_principal else opp.conta.telefone_principal,
            'email': opp.contato_principal.email if opp.contato_principal else opp.conta.email,
            'mensalidade': f"R$ {opp.valor_estimado:,.2f}".replace('.', 'X').replace(',', '.').replace('X', ','),
            'label_investimento': 'Investimento Anual' if opp.periodo_pagamento == 'ANUAL' else 'Mensalidade',
            'cupom': opp.cupom_desconto or "Nenhum",
            'forma_pagamento': opp.get_forma_pagamento_display(),
            'vendedor': opp.proprietario.get_full_name(),
            'indicador': opp.indicador_comissao.nome if opp.indicador_comissao else "Direto",
            'suporte': opp.canal.nome if opp.canal else "N/A"
        }
        
        template = f"""
Por gentileza, realizar o faturamento da contratação abaixo:

Plano: {params['plano_nome']} ({params['periodo']})
{params['recursos']}{params['adicionais']}

Cortesia:
• {params['cortesia']}

Dados do cliente responsável pelo fechamento:
• Nome: {params['cliente_nome']}
• WhatsApp: {params['whatsapp']}
• E-mail: {params['email']}

Investimento:
• {params['label_investimento']}: {params['mensalidade']}
• Cupom de desconto: {params['cupom']}

• Forma de pagamento: {params['forma_pagamento']}
• Vendedor: {params['vendedor']}
• Indicador da comissão: {params['indicador']}
"""
        if opp.canal:
            template += f"• Suporte: {params['suporte']}\n"
            
        template += "\nQualquer dúvida, estou a disposição!"
        
        return Response({'texto': template})

    @action(detail=True, methods=['get'])
    def historico_estagios(self, request, pk=None):
        """Retorna o histórico de mudanças de estágio da oportunidade"""
        from .models import HistoricoEstagio
        from .serializers import HistoricoEstagioSerializer
        
        oportunidade = self.get_object()
        historico = HistoricoEstagio.objects.filter(
            oportunidade=oportunidade
        ).select_related('usuario', 'estagio_anterior__estagio', 'estagio_novo__estagio')
        
        serializer = HistoricoEstagioSerializer(historico, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def gerar_proposta(self, request, pk=None):
        """Gera proposta comercial em HTML para a oportunidade"""
        from django.template import Template, Context
        from django.template.loader import render_to_string
        from django.http import HttpResponse
        
        opp = self.get_object()
        
        # Dados do plano
        plano = opp.plano
        if not plano:
            return Response({'error': 'Oportunidade não possui plano definido'}, status=400)
        
        # Calcula valor baseado no período
        if opp.periodo_pagamento == 'ANUAL' and plano.preco_anual:
            valor_mensal = f"{plano.preco_anual:,.0f}".replace(',', '.')
            periodo_texto = "Anual"
        else:
            valor_mensal = f"{plano.preco_mensal:,.0f}".replace(',', '.')
            periodo_texto = "Mensal"
        
        # Recursos do plano
        recursos = plano.recursos if plano.recursos else []
        
        # Adicionais contratados
        adicionais = []
        valor_adicionais = 0
        
        from .models import OportunidadeAdicional
        opp_adicionais = OportunidadeAdicional.objects.filter(oportunidade=opp).select_related('adicional')
        
        for oa in opp_adicionais:
            adicionais.append({
                'nome': oa.adicional.nome,
                'quantidade': oa.quantidade,
                'valor_unitario': float(oa.adicional.preco),
                'ativo': True
            })
            valor_adicionais += float(oa.adicional.preco) * oa.quantidade
        
        # Lista completa de adicionais (para mostrar todos, marcando os contratados)
        from .models import PlanoAdicional
        todos_adicionais = PlanoAdicional.objects.all()
        adicionais_lista = []
        
        for adicional in todos_adicionais:
            contratado = next((a for a in adicionais if a['nome'] == adicional.nome), None)
            adicionais_lista.append({
                'nome': adicional.nome,
                'quantidade': contratado['quantidade'] if contratado else 0,
                'ativo': contratado is not None
            })
        
        # Tabela de preços dos adicionais
        tabela_precos = []
        for adicional in todos_adicionais:
            tabela_precos.append({
                'nome': adicional.nome,
                'valor': f"{adicional.preco:,.0f}".replace(',', '.'),
                'unidade': 'por mês'
            })
        
        # Calcula valor total
        valor_plano = float(plano.preco_anual if opp.periodo_pagamento == 'ANUAL' and plano.preco_anual else plano.preco_mensal)
        valor_total = valor_plano + valor_adicionais
        valor_total_formatado = f"{valor_total:,.0f}".replace(',', '.')
        
        # Dados do vendedor
        vendedor = opp.proprietario
        vendedor_nome = vendedor.get_full_name() or vendedor.username
        vendedor_email = vendedor.email or ''
        vendedor_telefone = ''  # Adicionar campo se necessário
        
        # Renderiza o template
        context = {
            'plano_nome': plano.nome,
            'valor_mensal': valor_mensal,
            'periodo': periodo_texto,
            'recursos': recursos,
            'adicionais': adicionais,
            'adicionais_lista': adicionais_lista,
            'tabela_precos': tabela_precos,
            'valor_total': valor_total_formatado,
            'vendedor_nome': vendedor_nome,
            'vendedor_email': vendedor_email,
            'vendedor_telefone': vendedor_telefone,
            'conta_nome': opp.conta.nome_empresa if opp.conta else '',
            'oportunidade': opp,
        }
        
        try:
            html_content = render_to_string('proposta_comercial.html', context)
            
            # Retorna como HTML (frontend pode usar para preview ou converter em PDF)
            formato = request.query_params.get('formato', 'html')
            
            if formato == 'html':
                return HttpResponse(html_content, content_type='text/html')
            else:
                # Retorna dados JSON para o frontend renderizar
                return Response({
                    'html': html_content,
                    'dados': {
                        'plano_nome': plano.nome,
                        'valor_mensal': valor_mensal,
                        'periodo': periodo_texto,
                        'recursos': recursos,
                        'adicionais': adicionais,
                        'valor_total': valor_total_formatado,
                        'vendedor_nome': vendedor_nome,
                        'conta_nome': context['conta_nome'],
                    }
                })
        except Exception as e:
            logger.error(f"Erro ao gerar proposta: {str(e)}")
            return Response({'error': f'Erro ao gerar proposta: {str(e)}'}, status=500)


class OportunidadeAnexoViewSet(viewsets.ModelViewSet):
    """ViewSet para Anexos de Oportunidades"""
    queryset = OportunidadeAnexo.objects.all()
    serializer_class = OportunidadeAnexoSerializer
    permission_classes = [permissions.IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(uploaded_por=self.request.user)

    @action(detail=False, methods=['delete'])
    def batch_delete(self, request):
        ids = request.data.get('ids', [])
        OportunidadeAnexo.objects.filter(id__in=ids).delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class PlanoViewSet(viewsets.ModelViewSet):
    """ViewSet para CRUD de Planos (Admin cria/edita, todos leem)"""
    queryset = Plano.objects.all()
    serializer_class = PlanoSerializer
    
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [permissions.IsAuthenticated()]
        return [IsAdminUser()]


class PlanoAdicionalViewSet(viewsets.ModelViewSet):
    """ViewSet para CRUD de Adicionais de Plano (Admin cria/edita, todos leem)"""
    queryset = PlanoAdicional.objects.all()
    serializer_class = PlanoAdicionalSerializer
    
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [permissions.IsAuthenticated()]
        return [IsAdminUser()]


class ModuloTreinamentoViewSet(viewsets.ModelViewSet):
    """ViewSet para CRUD de Módulos de Treinamento (Admin cria/edita, todos leem)"""
    queryset = ModuloTreinamento.objects.all()
    serializer_class = ModuloTreinamentoSerializer

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [permissions.IsAuthenticated()]
        return [IsAdminUser()]


class OnboardingClienteViewSet(viewsets.ModelViewSet):
    """ViewSet para Fichas de Onboarding"""
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'responsavel', 'conta']
    search_fields = ['conta__nome_empresa']
    ordering_fields = ['data_criacao', 'data_inicio']

    def get_queryset(self):
        return OnboardingCliente.objects.select_related(
            'conta', 'responsavel', 'oportunidade'
        ).prefetch_related(
            'sessoes', 'sessoes__modulo', 'sessoes__treinador', 'sessoes__participantes'
        )

    def get_serializer_class(self):
        if self.action == 'list':
            return OnboardingClienteListSerializer
        return OnboardingClienteSerializer

    def perform_create(self, serializer):
        if not serializer.validated_data.get('responsavel'):
            serializer.save(responsavel=self.request.user)
        else:
            serializer.save()


class SessaoTreinamentoViewSet(viewsets.ModelViewSet):
    """ViewSet para Sessões de Treinamento individuais"""
    serializer_class = SessaoTreinamentoSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['onboarding', 'status', 'modulo']

    def get_queryset(self):
        return SessaoTreinamento.objects.select_related(
            'modulo', 'treinador', 'onboarding'
        ).prefetch_related('participantes')

    def perform_update(self, serializer):
        serializer.save()
        # Verifica se todas as sessões do onboarding estão concluídas
        sessao = serializer.instance
        onboarding = sessao.onboarding
        total = onboarding.sessoes.count()
        concluidas = onboarding.sessoes.filter(status='CONCLUIDO').count()
        if total > 0 and total == concluidas:
            onboarding.status = OnboardingCliente.STATUS_CONCLUIDO
            onboarding.save(update_fields=['status'])


class AgendaTreinamentoViewSet(viewsets.ModelViewSet):
    """ViewSet para Agendamentos de Treinamento (agenda)"""
    serializer_class = AgendaTreinamentoSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['onboarding', 'status', 'responsavel', 'modulo']
    search_fields = ['titulo', 'onboarding__conta__nome_empresa']
    ordering_fields = ['data', 'data_criacao']

    def get_queryset(self):
        qs = AgendaTreinamento.objects.select_related(
            'onboarding__conta', 'modulo', 'responsavel'
        )
        # Filtro por período
        data_inicio = self.request.query_params.get('data_inicio')
        data_fim = self.request.query_params.get('data_fim')
        if data_inicio:
            qs = qs.filter(data__gte=data_inicio)
        if data_fim:
            qs = qs.filter(data__lte=data_fim)
        # Filtro para próximos agendamentos
        proximos = self.request.query_params.get('proximos')
        if proximos:
            from datetime import date
            qs = qs.filter(data__gte=date.today(), status='AGENDADO')
        return qs

    def perform_create(self, serializer):
        if not serializer.validated_data.get('responsavel'):
            serializer.save(responsavel=self.request.user)
        else:
            serializer.save()


class AtividadeViewSet(viewsets.ModelViewSet):
    """ViewSet para Atividades"""
    serializer_class = AtividadeSerializer
    permission_classes = [HierarchyPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['tipo', 'status']
    search_fields = ['titulo', 'descricao']
    ordering_fields = ['data_vencimento', 'data_criacao']
    
    def get_queryset(self):
        user = self.request.user
        
        if user.perfil == 'ADMIN':
            return Atividade.objects.all()
        elif user.perfil == 'RESPONSAVEL':
            return Atividade.objects.filter(proprietario__canal=user.canal)
        else:
            return Atividade.objects.filter(proprietario=user)
    
    @action(detail=True, methods=['post'])
    def concluir(self, request, pk=None):
        """Marca uma atividade como concluída"""
        atividade = self.get_object()
        
        from django.utils import timezone
        atividade.status = Atividade.STATUS_CONCLUIDA
        if not atividade.data_conclusao:
            atividade.data_conclusao = timezone.now()
        atividade.save()
        
        serializer = self.get_serializer(atividade)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Retorna indicadores resumidos (pendentes e atrasadas)"""
        queryset = self.get_queryset()
        agora = timezone.now()
        
        total_pendentes = queryset.filter(status=Atividade.STATUS_PENDENTE).count()
        total_atrasadas = queryset.filter(
            status=Atividade.STATUS_PENDENTE, 
            data_vencimento__lt=agora
        ).count()
        
        return Response({
            'pendentes': total_pendentes,
            'atrasadas': total_atrasadas
        })

    @action(detail=False, methods=['get'], permission_classes=[permissions.AllowAny])
    def content_types(self, request):
        """Retorna os tipos de objetos que podem ser associados a atividades"""
        from django.contrib.contenttypes.models import ContentType
        models_permitidos = ['lead', 'conta', 'contato', 'oportunidade']
        cts = ContentType.objects.filter(
            app_label='crm', 
            model__in=models_permitidos
        )
        return Response([
            {'id': ct.id, 'model': ct.model, 'nome': ct.model.capitalize()} 
            for ct in cts
        ])


class DiagnosticoViewSet(viewsets.ModelViewSet):
    """ViewSet para o Diagnóstico de Maturidade"""
    queryset = DiagnosticoResultado.objects.all()
    serializer_class = DiagnosticoResultadoSerializer

    def get_permissions(self):
        """Define permissões: perguntas e submeter são públicos"""
        if self.action in ['perguntas', 'submeter', 'submeter_publico']:
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated(), HierarchyPermission()]

    def get_throttles(self):
        """Aplica rate limiting mais restrito nos endpoints públicos"""
        if self.action in ['perguntas', 'submeter', 'submeter_publico']:
            return [DiagnosticoPublicoThrottle()]
        return super().get_throttles()

    def get_queryset(self):
        """Filtra resultados do diagnóstico baseado na hierarquia (após autenticado)"""
        user = self.request.user
        if not user.is_authenticated:
            return DiagnosticoResultado.objects.none()
            
        if user.perfil == 'ADMIN':
            return DiagnosticoResultado.objects.all()
        elif user.perfil == 'RESPONSAVEL':
            return DiagnosticoResultado.objects.filter(lead__proprietario__canal=user.canal)
        else:
            return DiagnosticoResultado.objects.filter(lead__proprietario=user)

    @action(detail=False, methods=['get'])
    def perguntas(self, request):
        """Retorna todos os pilares, perguntas e respostas para o frontend público"""
        pilares = DiagnosticoPilar.objects.all().prefetch_related('perguntas__respostas')
        serializer = DiagnosticoPilarSerializer(pilares, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def submeter(self, request):
        """Recebe as respostas do diagnóstico, cria Contato/Oportunidade e salva o resultado"""
        serializer = DiagnosticoPublicSubmissionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        data = serializer.validated_data
        respostas_ids = data['respostas_ids']
        
        # 1. Busca as respostas no banco para calcular a pontuação
        respostas_objs = DiagnosticoResposta.objects.filter(
            id__in=respostas_ids
        ).select_related('pergunta__pilar')
        
        if not respostas_objs.exists():
            return Response(
                {'error': 'Nenhuma resposta válida encontrada'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 2. Processa pontuações por pilar
        pontuacao_por_pilar = {}
        respostas_detalhadas = []
        
        for r in respostas_objs:
            pilar_nome = r.pergunta.pilar.nome
            if pilar_nome not in pontuacao_por_pilar:
                pontuacao_por_pilar[pilar_nome] = {
                    'soma': 0,
                    'total_perguntas': 0,
                    'cor': r.pergunta.pilar.cor
                }
            
            pontuacao_por_pilar[pilar_nome]['soma'] += r.pontuacao
            pontuacao_por_pilar[pilar_nome]['total_perguntas'] += 1
            
            respostas_detalhadas.append({
                'pergunta': r.pergunta.texto,
                'pilar': pilar_nome,
                'resposta': r.texto,
                'pontos': r.pontuacao,
                'feedback': r.feedback
            })

        # Calcula a média (0-10) por pilar
        resultado_final = {}
        for pilar, dados in pontuacao_por_pilar.items():
            resultado_final[pilar] = {
                'score': round(dados['soma'] / dados['total_perguntas'], 1),
                'cor': dados['cor']
            }

        telefone_normalizado = self._normalize_phone_from_import(data.get('telefone', ''))

        try:
            with transaction.atomic():
                # 3. Cria ou busca Contato, Conta e Oportunidade
                admin_user = User.objects.filter(perfil='ADMIN', is_active=True).first()
                
                contato = Contato.objects.filter(email=data['email']).first()
                conta = contato.conta if contato else None
                oportunidade = None

                # Se veio oportunidade_id, tenta buscá-la
                if data.get('oportunidade_id'):
                    oportunidade = Oportunidade.objects.filter(id=data['oportunidade_id']).first()
                    if oportunidade:
                        contato = oportunidade.contato_principal
                        conta = oportunidade.conta

                # Se não temos contato mas temos contato_id
                if not contato and data.get('contato_id'):
                    contato = Contato.objects.filter(id=data['contato_id']).first()
                    if contato:
                        conta = contato.conta

                # Se ainda não temos contato por e-mail, cria um
                if not contato:
                    # Se temos empresa, tenta achar ou criar a conta
                    empresa_nome = data.get('empresa')
                    if empresa_nome and not conta:
                        conta = Conta.objects.filter(nome_empresa__iexact=empresa_nome).first()
                        if not conta:
                            conta = Conta.objects.create(
                                nome_empresa=empresa_nome,
                                proprietario=admin_user
                            )
                    
                    contato = Contato.objects.create(
                        nome=data['nome'],
                        email=data['email'],
                        telefone=telefone_normalizado or None,
                        conta=conta,
                        proprietario=admin_user
                    )

                # Se não temos oportunidade, cria uma nova
                if not oportunidade:
                    # Busca o funil "SDR" ou o primeiro do tipo OPORTUNIDADE
                    funil = Funil.objects.filter(nome__icontains='SDR').first() or \
                            Funil.objects.filter(tipo=Funil.TIPO_VENDAS).first() or \
                            Funil.objects.first()
                    
                    if funil:
                        vinculo_estagio = FunilEstagio.objects.filter(funil=funil, is_padrao=True).first() or \
                                         FunilEstagio.objects.filter(funil=funil).order_by('ordem').first()
                        
                        estagio = vinculo_estagio.estagio if vinculo_estagio else None
                        
                        oportunidade = Oportunidade.objects.create(
                            nome=f"Oportunidade - {contato.nome} (Diagnóstico)",
                            contato_principal=contato,
                            conta=conta,
                            funil=funil,
                            estagio=estagio,
                            proprietario=admin_user,
                            fonte='Diagnóstico de Maturidade'
                        )

                # 4. Salva o resultado do diagnóstico
                diagnostico = DiagnosticoResultado.objects.create(
                    conta=conta,
                    oportunidade=oportunidade,
                    respostas_detalhadas=respostas_detalhadas,
                    pontuacao_por_pilar=resultado_final
                )
                
                # 5. Gera Análise de IA
                diagnostico.analise_ia = gerar_analise_diagnostico(diagnostico)
                diagnostico.save()

                return Response({
                    'message': 'Diagnóstico processado com sucesso',
                    'contato_id': contato.id,
                    'oportunidade_id': oportunidade.id if oportunidade else None,
                    'resultado': resultado_final,
                    'analise_ia': diagnostico.analise_ia
                }, status=status.HTTP_201_CREATED)
                
        except Exception as e:
            logger.exception(f"Erro ao submeter diagnóstico: {e}")
            return Response(
                {'error': 'Erro ao processar diagnóstico. Tente novamente.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'], url_path='submeter-publico/(?P<canal_slug>[^/.]+)')
    def submeter_publico(self, request, canal_slug=None):
        """
        Endpoint PÚBLICO para submeter diagnóstico via link compartilhado pelo canal.
        
        URL: /api/diagnosticos/submeter-publico/<canal_slug>/
        
        O diagnóstico cria automaticamente:
        - Contato com os dados do cliente
        - Conta/Empresa
        - Oportunidade vinculada ao canal (ou matriz se canal não encontrado)
        """
        # 1. Busca o canal pelo slug (case-insensitive)
        canal = Canal.objects.filter(slug__iexact=canal_slug).first()
        
        # Define o proprietário: responsável do canal ou primeiro admin
        if canal and canal.responsavel:
            proprietario = canal.responsavel
        else:
            proprietario = User.objects.filter(perfil='ADMIN', is_active=True).first()
            canal = None  # Vai para matriz (sem canal)
        
        if not proprietario:
            return Response(
                {'error': 'Nenhum usuário disponível para receber a oportunidade'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        # 2. Valida os dados
        required_fields = ['nome', 'email', 'telefone', 'empresa', 'respostas_ids']
        for field in required_fields:
            if not request.data.get(field):
                return Response(
                    {'error': f'Campo obrigatório: {field}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        nome = request.data['nome']
        email = request.data['email']
        telefone = self._normalize_phone_from_import(request.data['telefone'])
        empresa = request.data['empresa']
        respostas_ids = request.data['respostas_ids']
        
        # Valida email
        from django.core.validators import validate_email
        from django.core.exceptions import ValidationError
        try:
            validate_email(email)
        except ValidationError:
            return Response(
                {'error': 'Email inválido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 3. Busca as respostas e calcula pontuação
        respostas_objs = DiagnosticoResposta.objects.filter(
            id__in=respostas_ids
        ).select_related('pergunta__pilar')
        
        if not respostas_objs.exists():
            return Response(
                {'error': 'Nenhuma resposta válida encontrada'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Processa pontuações por pilar
        pontuacao_por_pilar = {}
        respostas_detalhadas = []
        
        for r in respostas_objs:
            pilar_nome = r.pergunta.pilar.nome
            if pilar_nome not in pontuacao_por_pilar:
                pontuacao_por_pilar[pilar_nome] = {
                    'soma': 0,
                    'total_perguntas': 0,
                    'cor': r.pergunta.pilar.cor
                }
            
            pontuacao_por_pilar[pilar_nome]['soma'] += r.pontuacao
            pontuacao_por_pilar[pilar_nome]['total_perguntas'] += 1
            
            respostas_detalhadas.append({
                'pergunta': r.pergunta.texto,
                'pilar': pilar_nome,
                'resposta': r.texto,
                'pontos': r.pontuacao,
                'feedback': r.feedback
            })
        
        # Calcula a média (0-10) por pilar
        resultado_final = {}
        for pilar, dados in pontuacao_por_pilar.items():
            resultado_final[pilar] = {
                'score': round(dados['soma'] / dados['total_perguntas'], 1),
                'cor': dados['cor']
            }
        
        try:
            with transaction.atomic():
                # 4. Cria ou busca a Conta (Empresa)
                conta = Conta.objects.filter(nome_empresa__iexact=empresa).first()
                if not conta:
                    conta = Conta.objects.create(
                        nome_empresa=empresa,
                        canal=canal,
                        proprietario=proprietario
                    )
                
                # 5. Verifica se já existe contato com esse email
                contato = Contato.objects.filter(email__iexact=email).first()
                if not contato:
                    contato = Contato.objects.create(
                        nome=nome,
                        email=email,
                        telefone=telefone or None,
                        conta=conta,
                        canal=canal,
                        proprietario=proprietario
                    )
                else:
                    # Atualiza dados se necessário
                    if telefone and not contato.telefone:
                        contato.telefone = telefone
                        contato.save()
                
                # 6. Cria a Oportunidade
                # Usa funil/estágio do canal se configurado, senão usa padrão do sistema
                if canal and canal.funil_padrao:
                    funil = canal.funil_padrao
                    estagio = canal.estagio_inicial
                else:
                    # Fallback: primeiro funil ativo do tipo OPORTUNIDADE
                    funil = Funil.objects.filter(tipo=Funil.TIPO_VENDAS, is_active=True).first()
                    estagio = None
                
                if not funil:
                    return Response(
                        {'error': 'Nenhum funil de oportunidades configurado'},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR
                    )
                
                # Se não tem estágio definido, busca o padrão ou primeiro do funil
                if not estagio:
                    vinculo_estagio = FunilEstagio.objects.filter(funil=funil, is_padrao=True).first() or \
                                     FunilEstagio.objects.filter(funil=funil).order_by('ordem').first()
                    estagio = vinculo_estagio.estagio if vinculo_estagio else None
                
                oportunidade = Oportunidade.objects.create(
                    nome=f"Diagnóstico - {nome}",
                    contato_principal=contato,
                    conta=conta,
                    funil=funil,
                    estagio=estagio,
                    proprietario=proprietario,
                    fonte='Diagnóstico de Maturidade',
                    descricao=f"Oportunidade gerada via diagnóstico público.\nCanal: {canal.nome if canal else 'Matriz'}"
                )
                
                # 7. Salva o resultado do diagnóstico
                diagnostico = DiagnosticoResultado.objects.create(
                    conta=conta,
                    oportunidade=oportunidade,
                    respostas_detalhadas=respostas_detalhadas,
                    pontuacao_por_pilar=resultado_final
                )
                
                # 8. Gera Análise de IA (se disponível)
                try:
                    diagnostico.analise_ia = gerar_analise_diagnostico(diagnostico)
                    diagnostico.save()
                except Exception as e:
                    logger.error(f"Erro ao gerar análise IA: {e}")

                return Response({
                    'success': True,
                    'message': 'Diagnóstico enviado com sucesso!',
                    'oportunidade_id': oportunidade.id,
                    'resultado': resultado_final,
                    'analise_ia': diagnostico.analise_ia
                }, status=status.HTTP_201_CREATED)

        except Exception as e:
            logger.exception(f"Erro ao submeter diagnóstico público: {e}")
            return Response(
                {'error': 'Erro ao processar diagnóstico. Tente novamente.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class WhatsappViewSet(viewsets.ModelViewSet):
    """ViewSet para histórico e envio de mensagens WhatsApp"""
    queryset = WhatsappMessage.objects.all()
    serializer_class = WhatsappMessageSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['oportunidade', 'numero_remetente', 'numero_destinatario']
    ordering_fields = ['timestamp']
    pagination_class = ChatMessagesPagination

    def list(self, request, *args, **kwargs):
        """Usa serializer slim (sem media_base64) para evitar payloads enormes na listagem."""
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = WhatsappMessageSlimSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = WhatsappMessageSlimSerializer(queryset, many=True)
        return Response(serializer.data)

    def _get_evolution_service_for_entity(self, opp_id=None, canal_id=None):
        """
        Retorna um EvolutionService configurado para o canal correto.
        Prioridade: canal da Oportunidade > canal_id explícito > fallback global.

        Returns:
            tuple: (EvolutionService, Canal ou None, nome_instancia)
        """
        canal = None

        # 1) Tenta obter o canal da Oportunidade
        if opp_id:
            try:
                opp = Oportunidade.objects.select_related('canal').get(id=opp_id)
                canal = opp.canal
            except Oportunidade.DoesNotExist:
                pass

        # 2) Se não achou pelo opp, usa canal_id explícito (vem do multiatendimento)
        if not (canal and canal.evolution_instance_name) and canal_id:
            try:
                canal = Canal.objects.get(id=canal_id)
            except Canal.DoesNotExist:
                pass

        # Se encontrou canal com Evolution configurado, usa ele
        if canal and canal.evolution_instance_name:
            service = EvolutionService(
                instance_name=canal.evolution_instance_name,
                instance_token=canal.evolution_token
            )
            return service, canal, canal.evolution_instance_name

        # Fallback: usa instância global das settings
        service = EvolutionService()
        return service, None, service.instance

    def get_queryset(self):
        from django.conf import settings

        # Filtros de lead/oportunidade funcionam via DjangoFilterBackend
        # Mas o filtro de 'number' é customizado para pegar a conversa
        number = self.request.query_params.get('number')
        
        if number:
            # Limpa o número para pegar apenas dígitos
            clean_number = ''.join(filter(str.isdigit, str(number)))
            
            # Criamos variações do número para aumentar a chance de match
            # Ex: com 55, sem 55, com 9o dígito, sem 9o dígito
            variations = [clean_number]
            
            # Se for um número brasileiro sem 55
            if len(clean_number) >= 10 and not clean_number.startswith('55'):
                variations.append('55' + clean_number)
            
            # Se for um número brasileiro com 55
            if clean_number.startswith('55') and len(clean_number) >= 12:
                variations.append(clean_number[2:]) # Versão sem DDI
                
                # Lidar com o 9o dígito (específico do Brasil)
                # Formato: 55 + DDD (2) + 9 + Numero (8) = 13 dígitos
                if len(clean_number) == 13:
                    # Versão sem o 9
                    variations.append(clean_number[:4] + clean_number[5:])
                # Formato: 55 + DDD (2) + Numero (8) = 12 dígitos
                elif len(clean_number) == 12:
                    # Versão com o 9
                    variations.append(clean_number[:4] + '9' + clean_number[4:])
            
            # Remove duplicatas
            variations = list(set(variations))
            
            # Busca mensagens onde qualquer uma das variações aparece em remetente ou destinatário
            q_filter = Q()
            for v in variations:
                q_filter |= Q(numero_remetente__icontains=v)
                q_filter |= Q(numero_destinatario__icontains=v)
            
            # Limita período para evitar carregar histórico inteiro (padrão: 90 dias)
            dias_chat_raw = self.request.query_params.get('dias', 90)
            try:
                dias_chat = int(dias_chat_raw)
            except (TypeError, ValueError):
                dias_chat = 90
            data_limite = timezone.now() - timedelta(days=dias_chat)
            
            queryset = self.queryset.filter(q_filter).filter(timestamp__gte=data_limite)
            
            # Exclui números bloqueados
            bloqueados = list(NumeroBloqueado.objects.values_list('numero', flat=True))
            if bloqueados:
                q_block = Q()
                for nb in bloqueados:
                    q_block |= Q(numero_remetente__icontains=nb)
                    q_block |= Q(numero_destinatario__icontains=nb)
                queryset = queryset.exclude(q_block)
            
            # Compatibilidade: cliente antigo pode enviar "limit" sem paginação.
            has_page = self.request.query_params.get('page') is not None
            has_page_size = self.request.query_params.get('page_size') is not None
            limit_raw = self.request.query_params.get('limit')
            if limit_raw and not (has_page or has_page_size):
                try:
                    max_msgs = int(limit_raw)
                except (TypeError, ValueError):
                    max_msgs = 100
                if max_msgs > 0:
                    ids_recentes = queryset.order_by('-timestamp').values_list('id', flat=True)[:max_msgs]
                    queryset = queryset.filter(id__in=ids_recentes)
            
            return queryset.order_by('timestamp')

        return super().get_queryset()

    # ==================== ENDPOINTS DE CONEXÃO ====================

    @action(detail=False, methods=['get'])
    def status(self, request):
        """Retorna o status da conexão WhatsApp"""
        service = EvolutionService()
        result = service.get_connection_status()
        return Response(result)

    @action(detail=False, methods=['get'])
    def qrcode(self, request):
        """Retorna o QR Code para conexão"""
        service = EvolutionService()
        result = service.get_qr_code()
        return Response(result)

    @action(detail=False, methods=['post'])
    def connect(self, request):
        """Inicia o processo de conexão e retorna QR Code"""
        service = EvolutionService()
        
        # Primeiro verifica o status atual
        status_result = service.get_connection_status()
        
        if status_result.get('connected'):
            return Response({
                'success': True,
                'already_connected': True,
                'message': 'WhatsApp já está conectado',
                'status': status_result
            })
        
        # Se não está conectado, obtém o QR Code
        qr_result = service.get_qr_code()
        
        return Response({
            'success': qr_result.get('success', False),
            'already_connected': False,
            'qr_code': qr_result.get('qr_code'),
            'qr_base64': qr_result.get('qr_base64'),
            'status': status_result
        })

    @action(detail=False, methods=['post'])
    def disconnect(self, request):
        """Desconecta o WhatsApp"""
        service = EvolutionService()
        result = service.disconnect()
        return Response(result)

    @action(detail=False, methods=['post'])
    def restart(self, request):
        """Reinicia a instância"""
        service = EvolutionService()
        result = service.restart_instance()
        return Response(result)

    @action(detail=False, methods=['get'])
    def instance_info(self, request):
        """Retorna informações da instância"""
        service = EvolutionService()
        result = service.get_instance_info()
        return Response(result)

    # ==================== ENDPOINTS DE MENSAGEM ====================


    @action(detail=False, methods=['post'])
    def send(self, request):
        """Action para enviar mensagem através do CRM"""
        import uuid

        number = request.data.get('number')
        text = request.data.get('text')
        opp_id = request.data.get('oportunidade')
        canal_id = request.data.get('canal_id')

        if not number or not text:
            return Response({'error': 'Número e texto são obrigatórios'}, status=400)

        # Obtém o serviço Evolution do canal correto
        service, canal, instance_name = self._get_evolution_service_for_entity(opp_id, canal_id)

        try:
            result = service.send_text(number, text)

            # Extrai ID da mensagem da resposta da Evolution API
            # A estrutura pode variar bastante, vamos tentar todos os formatos possíveis
            msg_id = None
            if isinstance(result, dict):
                # Tenta formato direto: { key: { id: "..." } }
                if 'key' in result and isinstance(result['key'], dict):
                    msg_id = result['key'].get('id')
                # Tenta formato aninhado: { data: { key: { id: "..." } } }
                elif 'data' in result and isinstance(result['data'], dict):
                    data_obj = result['data']
                    if 'key' in data_obj and isinstance(data_obj['key'], dict):
                        msg_id = data_obj['key'].get('id')
                # Tenta formato alternativo: { message: { key: { id: "..." } } }
                elif 'message' in result and isinstance(result['message'], dict):
                    if 'key' in result['message'] and isinstance(result['message']['key'], dict):
                        msg_id = result['message']['key'].get('id')

            # Se não conseguiu extrair, gera um ID único local
            if not msg_id:
                msg_id = f"local_{uuid.uuid4().hex[:20]}"
                logger.debug(f"[SEND] ID não encontrado na resposta Evolution, usando ID local: {msg_id}")

            # Formata o número para armazenamento consistente
            formatted_number = service._format_number(number)

            # Salva localmente (webhook pode ter chegado antes — usa get_or_create para evitar IntegrityError)
            msg, _ = WhatsappMessage.objects.get_or_create(
                id_mensagem=msg_id,
                defaults=dict(
                    instancia=instance_name,
                    de_mim=True,
                    numero_remetente=instance_name,
                    numero_destinatario=formatted_number,
                    texto=text,
                    timestamp=timezone.now(),
                    oportunidade_id=opp_id
                )
            )

            return Response(WhatsappMessageSerializer(msg).data)
        except Exception as e:
            logger.exception(f"Erro ao enviar mensagem WhatsApp: {e}")
            return Response({'error': str(e)}, status=500)

    @action(detail=False, methods=['post'])
    def send_media(self, request):
        """Action para enviar mídia (imagem) através do CRM"""
        import uuid

        number = request.data.get('number')
        media = request.data.get('media')  # Base64
        media_type = request.data.get('mediaType', 'image')
        file_name = request.data.get('fileName', 'image.jpg')
        caption = request.data.get('caption', '')
        opp_id = request.data.get('oportunidade')
        canal_id = request.data.get('canal_id')

        if not number or not media:
            return Response({'error': 'Número e mídia são obrigatórios'}, status=400)

        # Obtém o serviço Evolution do canal correto
        service, canal, instance_name = self._get_evolution_service_for_entity(opp_id, canal_id)

        try:
            # Remove prefixo data:image/xxx;base64, se existir
            if ';base64,' in media:
                media = media.split(';base64,')[1]

            result = service.send_media(number, media, media_type, file_name, caption)

            # Extrai ID da mensagem
            msg_id = None
            if isinstance(result, dict):
                if 'key' in result and isinstance(result['key'], dict):
                    msg_id = result['key'].get('id')
                elif 'data' in result and isinstance(result['data'], dict):
                    data_obj = result['data']
                    if 'key' in data_obj and isinstance(data_obj['key'], dict):
                        msg_id = data_obj['key'].get('id')

            if not msg_id:
                msg_id = f"local_{uuid.uuid4().hex[:20]}"
            
            formatted_number = service._format_number(number)
            
            # Determina o mimetype baseado no nome do arquivo
            import mimetypes
            mime_type, _ = mimetypes.guess_type(file_name)
            if not mime_type:
                mime_type = 'image/jpeg'
            
            # Salva localmente com base64 completo (webhook pode ter chegado antes)
            media_b64 = f"data:{mime_type};base64,{media}" if media_type == 'image' else None
            msg, created = WhatsappMessage.objects.get_or_create(
                id_mensagem=msg_id,
                defaults=dict(
                    instancia=instance_name,
                    de_mim=True,
                    numero_remetente=instance_name,
                    numero_destinatario=formatted_number,
                    texto=caption or f"📷 [Imagem: {file_name}]",
                    tipo_mensagem=media_type,
                    media_base64=media_b64,
                    timestamp=timezone.now(),
                    oportunidade_id=opp_id
                )
            )
            # Se o webhook criou antes (sem media_base64), atualiza com o dado local
            if not created and media_b64 and not msg.media_base64:
                msg.media_base64 = media_b64
                msg.save(update_fields=['media_base64'])

            return Response(WhatsappMessageSerializer(msg).data)
        except Exception as e:
            logger.exception(f"Erro ao enviar mídia WhatsApp: {e}")
            return Response({'error': str(e)}, status=500)

    @action(detail=False, methods=['post'])
    def sync(self, request):
        """
        Sincroniza mensagens da Evolution API para um número específico.
        Busca as mensagens mais recentes da API e importa as que ainda não existem.
        """
        number = request.data.get('number')
        lead_id = request.data.get('lead')
        opp_id = request.data.get('oportunidade')
        canal_id = request.data.get('canal_id')
        limit = request.data.get('limit', 50)
        dias_sync = int(request.data.get('dias', 30))

        if not number:
            return Response({'error': 'Número é obrigatório'}, status=400)

        # Obtém o serviço Evolution do canal correto
        service, canal, instance_name = self._get_evolution_service_for_entity(opp_id, canal_id)

        # Limita busca aos últimos N dias para evitar puxar histórico inteiro
        ts_start = int((timezone.now() - timedelta(days=dias_sync)).timestamp())

        try:
            # Busca mensagens da API Evolution (limitadas por período)
            api_messages = service.find_messages(number, limit=limit, timestamp_start=ts_start)
            
            # Se a resposta for um dict com 'messages', extrai a lista
            if isinstance(api_messages, dict):
                api_messages = api_messages.get('messages', api_messages.get('data', []))
            
            if not isinstance(api_messages, list):
                api_messages = []
            
            imported_count = 0
            skipped_count = 0
            
            for msg_data in api_messages:
                key = msg_data.get('key', {})
                id_msg = key.get('id')
                
                if not id_msg:
                    continue
                
                # Verifica se já existe
                if WhatsappMessage.objects.filter(id_mensagem=id_msg).exists():
                    skipped_count += 1
                    continue
                
                remote_jid = key.get('remoteJid', '')
                from_me = key.get('fromMe', False)
                
                # Extrai texto
                message_content = msg_data.get('message', {})
                text = ""
                if 'conversation' in message_content:
                    text = message_content['conversation']
                elif 'extendedTextMessage' in message_content:
                    text = message_content['extendedTextMessage'].get('text', '')
                
                # Mídia
                mtype = 'text'
                if not text:
                    for media_type in ['imageMessage', 'videoMessage', 'documentMessage', 'audioMessage']:
                        if media_type in message_content:
                            text = message_content[media_type].get('caption', f'[{media_type}]')
                            mtype = media_type.replace('Message', '')
                            break
                
                # Timestamp
                ts_int = msg_data.get('messageTimestamp')
                if ts_int:
                    dt = timezone.datetime.fromtimestamp(int(ts_int), tz=dt_timezone.utc)
                else:
                    dt = timezone.now()
                
                # Determina remetente/destinatário usando a instância correta
                remote_number = remote_jid.split('@')[0] if remote_jid else ''
                
                if from_me:
                    numero_remetente = instance_name
                    numero_destinatario = remote_number
                else:
                    numero_remetente = remote_number
                    numero_destinatario = instance_name
                
                # Cria a mensagem
                msg_obj = WhatsappMessage.objects.create(
                    id_mensagem=id_msg,
                    instancia=instance_name,
                    de_mim=from_me,
                    numero_remetente=numero_remetente,
                    numero_destinatario=numero_destinatario,
                    texto=text or '[sem texto]',
                    tipo_mensagem=mtype,
                    timestamp=dt,
                    oportunidade_id=opp_id
                )

                # Tenta linkar com Oportunidade se não foi fornecido
                if not opp_id:
                    EvolutionService.identify_and_link_message(msg_obj)
                
                imported_count += 1
            
            return Response({
                'imported': imported_count,
                'skipped': skipped_count,
                'message': f'{imported_count} mensagens importadas, {skipped_count} já existiam'
            })

        except Exception as e:
            logger.exception(f"Erro ao sincronizar mensagens WhatsApp: {e}")
            return Response({'error': str(e)}, status=500)

    @action(detail=False, methods=['post'])
    def marcar_lidas(self, request):
        """Marca mensagens como lidas para um número, lead ou oportunidade"""
        number = request.data.get('number')
        lead_id = request.data.get('lead')
        opp_id = request.data.get('oportunidade')
        
        # Filtro base: mensagens recebidas e não lidas
        q_filter = Q(lida=False, de_mim=False)
        
        # Se informou ID, prioriza ele, mas se não achar nada pelo ID, tenta pelo número
        # Isso ajuda se a vinculação falhou
        if lead_id or opp_id or number:
            sub_filter = Q()
            if lead_id:
                sub_filter |= Q(lead_id=lead_id)
            if opp_id:
                sub_filter |= Q(oportunidade_id=opp_id)
            if number:
                clean_number = ''.join(filter(str.isdigit, str(number)))
                if len(clean_number) >= 8:
                    sub_filter |= Q(numero_remetente__icontains=clean_number)
            
            q_filter &= sub_filter
        else:
            return Response({'error': 'Informe number, lead ou oportunidade'}, status=400)
            
        updated = WhatsappMessage.objects.filter(q_filter).update(lida=True)
        return Response({'status': 'success', 'updated_count': updated})

    @action(detail=False, methods=['get'])
    def unread_counts(self, request):
        """Retorna o total de mensagens não lidas para o usuário logado"""
        user = request.user
        
        # Filtro base: mensagens recebidas e não lidas
        unread_base = WhatsappMessage.objects.filter(lida=False, de_mim=False)
        
        # Filtros de Hierarquia
        if user.perfil == 'ADMIN':
            # ADMIN vê tudo
            outros_unread = unread_base.filter(oportunidade__isnull=True).count()
            opps_unread = unread_base.filter(oportunidade__isnull=False).count()
        elif user.perfil == 'RESPONSAVEL':
            # RESPONSAVEL vê do seu canal
            outros_unread = unread_base.filter(
                oportunidade__isnull=True
            ).count() # TODO: Filtrar outros por canal se possível via instância
            opps_unread = unread_base.filter(
                Q(oportunidade__canal=user.canal) | Q(oportunidade__proprietario__canal=user.canal)
            ).count()
        else: # VENDEDOR
            # VENDEDOR vê apenas o que é dele
            outros_unread = 0 # Vendedor geralmente não vê novos que não são dele
            opps_unread = unread_base.filter(oportunidade__proprietario=user).count()
            
        return Response({
            'novas': outros_unread,
            'oportunidades': opps_unread,
            'total': outros_unread + opps_unread
        })

    @action(detail=False, methods=['post'])
    def process_pending_media(self, request):
        """
        Processa mídias pendentes (áudios não transcritos, imagens sem base64).
        Chamado quando o chat é aberto para garantir que as mídias sejam processadas.
        """
        number = request.data.get('number')
        if not number:
            return Response({'error': 'number required'}, status=400)
        
        # Remove formatação do número
        import re
        clean_number = re.sub(r'\D', '', str(number))
        
        # Gera variações do número (com e sem 9º dígito)
        variations = set([clean_number])
        
        # Remove DDI se existir
        base = clean_number[2:] if clean_number.startswith('55') else clean_number
        variations.add(base)
        variations.add('55' + base)
        
        # Variação com/sem 9º dígito
        if len(base) == 11 and base[2] == '9':
            # Tem 9, gera sem
            without_9 = base[:2] + base[3:]
            variations.add(without_9)
            variations.add('55' + without_9)
        elif len(base) == 10:
            # Não tem 9, gera com
            with_9 = base[:2] + '9' + base[2:]
            variations.add(with_9)
            variations.add('55' + with_9)
        
        # Adiciona últimos 8 dígitos
        if len(clean_number) >= 8:
            variations.add(clean_number[-8:])
        
        # Busca mensagens pendentes com todas as variações
        from django.db.models import Q
        
        q_filter = Q()
        for v in variations:
            if len(v) >= 8:
                q_filter |= Q(numero_remetente__icontains=v) | Q(numero_destinatario__icontains=v)
        
        pending_audio = WhatsappMessage.objects.filter(
            q_filter,
            tipo_mensagem='audio',
            texto__in=['🎤 [Áudio]', '🎤 [Áudio não transcrito]', '[audioMessage]']
        )
        
        pending_images = WhatsappMessage.objects.filter(
            q_filter,
            tipo_mensagem='image',
            media_base64__isnull=True
        )
        
        processed_audio = 0
        processed_images = 0
        
        # Processa áudios
        if pending_audio.exists():
            from .services.evolution_api import EvolutionService
            from .services.audio_transcription import transcribe_from_base64
            
            for msg in pending_audio[:5]:  # Limita para não demorar muito
                try:
                    # Obtém a instância Evolution do canal da mensagem
                    service, canal, instance_name = self._get_evolution_service_for_entity(msg.oportunidade_id)
                    
                    key = {
                        'id': msg.id_mensagem,
                        'remoteJid': f"{msg.numero_remetente}@s.whatsapp.net",
                        'fromMe': msg.de_mim
                    }
                    
                    media_result = service.get_media_base64(key)
                    
                    if media_result and media_result.get('base64'):
                        # Salva base64 do áudio para reprodução futura
                        mimetype = media_result.get('mimetype', 'audio/ogg; codecs=opus')
                        audio_b64 = media_result['base64']
                        if not audio_b64.startswith('data:'):
                            audio_b64 = f"data:{mimetype};base64,{audio_b64}"
                        msg.media_base64 = audio_b64

                        transcription = transcribe_from_base64(
                            media_result['base64'],
                            media_result.get('mimetype', '')
                        )
                        
                        if transcription and transcription.get('text'):
                            duration = transcription.get('duration', 0)
                            msg.texto = f"🎤 [Áudio {int(duration)}s]: {transcription['text']}"
                            msg.save(update_fields=['texto', 'media_base64'])
                            processed_audio += 1
                        else:
                            msg.save(update_fields=['media_base64'])
                            
                except Exception as e:
                    logger.error(f"[ProcessMedia] Erro ao processar áudio {msg.id}: {e}")

        # Processa imagens
        if pending_images.exists():
            from .services.evolution_api import EvolutionService
            
            for msg in pending_images[:10]:  # Limita
                try:
                    # Obtém a instância Evolution do canal da mensagem
                    service, canal, instance_name = self._get_evolution_service_for_entity(msg.oportunidade_id)
                    
                    key = {
                        'id': msg.id_mensagem,
                        'remoteJid': f"{msg.numero_remetente}@s.whatsapp.net",
                        'fromMe': msg.de_mim
                    }
                    
                    media_result = service.get_media_base64(key)
                    
                    if media_result and media_result.get('base64'):
                        mimetype = media_result.get('mimetype', 'image/jpeg')
                        base64_data = media_result['base64']
                        
                        if not base64_data.startswith('data:'):
                            base64_data = f"data:{mimetype};base64,{base64_data}"
                        
                        msg.media_base64 = base64_data
                        msg.save(update_fields=['media_base64'])
                        processed_images += 1
                        
                except Exception as e:
                    logger.error(f"[ProcessMedia] Erro ao processar imagem {msg.id}: {e}")
        
        return Response({
            'processed_audio': processed_audio,
            'processed_images': processed_images
        })

    @action(detail=False, methods=['post'])
    def get_audio(self, request):
        """
        Baixa apenas o áudio de uma mensagem sem transcrever.
        Útil para reprodução sem processamento.
        """
        message_id = request.data.get('message_id')

        if not message_id:
            return Response({'error': 'message_id required'}, status=400)

        try:
            msg = WhatsappMessage.objects.get(id=message_id)
        except WhatsappMessage.DoesNotExist:
            return Response({'error': 'message not found'}, status=404)

        if msg.tipo_mensagem != 'audio':
            return Response({'error': 'message is not audio'}, status=400)

        # Verifica se já tem o áudio salvo no banco (cache)
        if msg.media_base64:
            logger.info(f"[GetAudio] msg={msg.id} servido do cache DB")
            return Response({
                'success': True,
                'audio_url': msg.media_base64,
                'mimetype': 'audio/ogg; codecs=opus'
            })

        # Baixa o áudio da Evolution API usando a instância correta
        from .services.evolution_api import EvolutionService

        # Obtém a instância Evolution do canal associado
        service, canal, instance_name = self._get_evolution_service_for_entity(msg.oportunidade_id)

        # Usa o remoteJid correto: para recebidas é o número do remetente, para enviadas é o destinatário
        remote_number = msg.numero_remetente if not msg.de_mim else msg.numero_destinatario
        # Remove @s.whatsapp.net se já vier com sufixo
        remote_number = remote_number.split('@')[0] if remote_number else ''
        key = {
            'id': msg.id_mensagem,
            'remoteJid': f"{remote_number}@s.whatsapp.net",
            'fromMe': msg.de_mim
        }

        logger.info(f"[GetAudio] msg={msg.id} key={key}")
        media_result = service.get_media_base64(key)

        if not media_result or not media_result.get('base64'):
            logger.error(f"[GetAudio] Falha ao baixar áudio para msg={msg.id} key={key}")
            return Response({'error': 'could_not_download_audio'}, status=500)

        base64_data = media_result['base64']
        mimetype = media_result.get('mimetype') or 'audio/ogg; codecs=opus'
        if not mimetype.startswith('audio'):
            mimetype = 'audio/ogg; codecs=opus'

        audio_url = f"data:{mimetype};base64,{base64_data}"

        # Salva no banco para próximas requisições
        try:
            msg.media_base64 = audio_url
            msg.save(update_fields=['media_base64'])
            logger.info(f"[GetAudio] msg={msg.id} áudio salvo no DB para cache")
        except Exception as e:
            logger.warning(f"[GetAudio] Falha ao salvar cache DB: {e}")

        return Response({
            'success': True,
            'audio_url': audio_url,
            'mimetype': mimetype
        })

    @action(detail=False, methods=['post'])
    def transcribe_audio(self, request):
        """
        Transcreve um áudio específico por ID.
        Retorna a transcrição e o base64 do áudio para reprodução.
        """
        message_id = request.data.get('message_id')

        if not message_id:
            return Response({'error': 'message_id required'}, status=400)

        try:
            msg = WhatsappMessage.objects.get(id=message_id)
        except WhatsappMessage.DoesNotExist:
            return Response({'error': 'message not found'}, status=404)

        if msg.tipo_mensagem != 'audio':
            return Response({'error': 'message is not audio'}, status=400)

        from .services.audio_transcription import transcribe_from_base64

        base64_data = None
        mimetype = 'audio/ogg'
        audio_url = None

        # Verifica se já tem o áudio salvo no banco (cache)
        if msg.media_base64:
            logger.info(f"[TranscribeAudio] msg={msg.id} usando áudio do cache DB")
            audio_url = msg.media_base64
            # Extrai base64 puro e mimetype do data URI para transcrição
            if msg.media_base64.startswith('data:'):
                parts = msg.media_base64.split(',', 1)
                if len(parts) == 2:
                    mimetype = parts[0].replace('data:', '').replace(';base64', '')
                    base64_data = parts[1]
            else:
                base64_data = msg.media_base64

        # Se não tem no cache, baixa da Evolution API
        if not base64_data:
            from .services.evolution_api import EvolutionService

            service, canal, instance_name = self._get_evolution_service_for_entity(msg.oportunidade_id)
            remote_number = msg.numero_remetente if not msg.de_mim else msg.numero_destinatario
            remote_number = remote_number.split('@')[0] if remote_number else ''
            key = {
                'id': msg.id_mensagem,
                'remoteJid': f"{remote_number}@s.whatsapp.net",
                'fromMe': msg.de_mim
            }

            media_result = service.get_media_base64(key)

            if not media_result or not media_result.get('base64'):
                return Response({'error': 'could_not_download_audio'}, status=500)

            base64_data = media_result['base64']
            mimetype = media_result.get('mimetype', 'audio/ogg')

            # Formata o base64 para reprodução
            audio_url = f"data:{mimetype};base64,{base64_data}"

            # Salva no banco para próximas requisições
            try:
                msg.media_base64 = audio_url
                msg.save(update_fields=['media_base64'])
                logger.info(f"[TranscribeAudio] msg={msg.id} áudio salvo no DB para cache")
            except Exception as e:
                logger.warning(f"[TranscribeAudio] Falha ao salvar cache DB: {e}")

        # Tenta transcrever
        transcription_text = None
        duration = 0
        transcription_error = None

        try:
            logger.info(f"[TranscribeAudio] Iniciando transcrição para msg {message_id}")
            logger.info(f"[TranscribeAudio] Mimetype: {mimetype}, Base64 len: {len(base64_data)}")

            result = transcribe_from_base64(base64_data, mimetype)
            logger.info(f"[TranscribeAudio] Resultado: {result}")

            if result and result.get('text'):
                transcription_text = result['text']
                duration = result.get('duration', 0)

                # Atualiza a mensagem no banco
                msg.texto = f"🎤 [Áudio {int(duration)}s]: {transcription_text}"
                msg.save(update_fields=['texto'])
                logger.info(f"[TranscribeAudio] Sucesso! Texto: {transcription_text[:50]}...")
            else:
                transcription_error = "transcription_empty"
                logger.warning(f"[TranscribeAudio] {transcription_error}")
        except Exception as e:
            transcription_error = str(e)
            logger.error(f"[TranscribeAudio] Erro: {e}")
            import traceback
            logger.error(traceback.format_exc())

        return Response({
            'success': True,
            'audio_url': audio_url,
            'transcription': transcription_text,
            'duration': duration,
            'updated_text': msg.texto,
            'error': transcription_error
        })


class WhatsappWebhookView(APIView):
    """Recebe notificações da Evolution API (MESSAGES_UPSERT)"""
    permission_classes = [permissions.AllowAny]  # Evolution envia sem auth JWT
    throttle_classes = [WebhookThrottle]

    def _validate_webhook_token(self, request):
        """Valida o token secreto do webhook se WEBHOOK_SECRET estiver configurado."""
        webhook_secret = getattr(settings, 'WEBHOOK_SECRET', None) or ''
        if not webhook_secret:
            # Se não configurado, aceita sem validação (compatibilidade)
            return True
        incoming_token = request.headers.get('apikey', '') or request.headers.get('Authorization', '')
        return incoming_token == webhook_secret

    def post(self, request):
        import json

        if not self._validate_webhook_token(request):
            logger.warning("[WEBHOOK] Requisição rejeitada: token inválido")
            return Response({'error': 'Unauthorized'}, status=401)

        data = request.data

        # Log para debug (ajuda muito a identificar mudanças na API Evolution)
        logger.warning(f"[WEBHOOK] Payload recebido | event={data.get('event')} instance={data.get('instance')} | {json.dumps(data)[:800]}")
        
        event = data.get('event', '').lower()
        instance = data.get('instance', 'unknown')

        # Normaliza o nome do evento
        event = event.replace('_', '.')
        
        # Aceita mensagens recebidas/atribuídas e enviadas
        if 'messages' in event or 'message' in event:
            # Tenta encontrar a lista de mensagens em qualquer lugar do payload
            messages = []
            
            # Formatos comuns da Evolution API:
            # 1. data['data']['messages']
            # 2. data['messages']
            # 3. data['data'] (se for uma única mensagem)
            
            if 'data' in data and isinstance(data['data'], dict) and 'messages' in data['data']:
                messages = data['data']['messages']
            elif 'messages' in data:
                messages = data['messages']
            elif 'data' in data:
                messages = [data['data']] if isinstance(data['data'], dict) else []
            else:
                # Se não achou nos lugares comuns, procura por 'key' no root (formado de mensagem única)
                if 'key' in data:
                    messages = [data]

            if not messages:
                # print(f"[WEBHOOK] Nenhum conteúdo de mensagem encontrado para o evento {event}", file=sys.stderr)
                return Response({'status': 'ignored'}, status=200)

            # print(f"[WEBHOOK] Processando {len(messages)} mensagens do evento '{event}'", file=sys.stderr)

            for msg_data in messages:
                try:
                    key = msg_data.get('key', {})
                    id_msg = key.get('id')

                    if not id_msg:
                        continue
                    
                    remote_jid = key.get('remoteJid', '')
                    from_me = key.get('fromMe', False)
                    
                    # Trata reações (reactionMessage) — atualiza a mensagem original
                    message_content = msg_data.get('message', {})
                    if 'reactionMessage' in message_content:
                        reaction = message_content['reactionMessage']
                        original_id = reaction.get('key', {}).get('id')
                        emoji = reaction.get('text', '')
                        remote_number = remote_jid.split('@')[0] if remote_jid else ''
                        try:
                            original_msg = WhatsappMessage.objects.get(id_mensagem=original_id)
                            reacoes = list(original_msg.reacoes or [])
                            # Remove reação anterior do mesmo número
                            reacoes = [r for r in reacoes if r.get('numero') != remote_number]
                            if emoji:  # string vazia = remoção da reação
                                reacoes.append({'emoji': emoji, 'de_mim': from_me, 'numero': remote_number})
                            original_msg.reacoes = reacoes
                            original_msg.save(update_fields=['reacoes'])
                        except WhatsappMessage.DoesNotExist:
                            pass
                        continue  # Reação não é salva como mensagem nova

                    # Previne duplicatas
                    if WhatsappMessage.objects.filter(id_mensagem=id_msg).exists():
                        continue

                    # Extrai texto
                    message_content = msg_data.get('message', {})
                    text = ""
                    media_url = None
                    
                    if 'conversation' in message_content:
                        text = message_content['conversation']
                    elif 'extendedTextMessage' in message_content:
                        text = message_content['extendedTextMessage'].get('text', '')
                    elif 'buttonsResponseMessage' in message_content:
                        text = message_content['buttonsResponseMessage'].get('selectedDisplayText', '')
                    
                    # Mídia
                    mtype = 'text'
                    media_base64 = None
                    needs_async_processing = False
                    
                    if not text:
                        for media_type in ['imageMessage', 'videoMessage', 'documentMessage', 'audioMessage']:
                            if media_type in message_content:
                                media_content = message_content[media_type]
                                text = media_content.get('caption', '')
                                mtype = media_type.replace('Message', '')
                                
                                # Extrai URL da mídia para referência
                                media_url = media_content.get('url') or media_content.get('directPath')

                                # Captura base64 se a Evolution API enviou via Webhook Base64
                                inline_b64 = media_content.get('base64') or msg_data.get('base64')
                                if inline_b64:
                                    mimetype = media_content.get('mimetype') or media_content.get('mimeType') or ''
                                    if not inline_b64.startswith('data:'):
                                        inline_b64 = f"data:{mimetype};base64,{inline_b64}" if mimetype else inline_b64
                                    media_base64 = inline_b64

                                # Define texto temporário; só marca async se não tiver base64 inline
                                if media_type == 'audioMessage':
                                    if not text:
                                        text = "🎤 [Áudio]"
                                    if not media_base64:
                                        needs_async_processing = True

                                elif media_type == 'imageMessage':
                                    if not text:
                                        text = "📷 [Imagem]"
                                    if not media_base64:
                                        needs_async_processing = True
                                    
                                elif media_type == 'videoMessage':
                                    if not text:
                                        text = "🎥 [Vídeo]"
                                        
                                elif media_type == 'documentMessage':
                                    filename = media_content.get('fileName', 'documento')
                                    if not text:
                                        text = f"📄 [{filename}]"
                                
                                if not text:
                                    text = f'[{media_type}]'
                                break

                    # Timestamp
                    ts_int = msg_data.get('messageTimestamp')
                    dt = timezone.datetime.fromtimestamp(int(ts_int), tz=dt_timezone.utc) if ts_int else timezone.now()

                    # Determina números
                    remote_number = remote_jid.split('@')[0] if remote_jid else ''
                    
                    if from_me:
                        numero_remetente = instance
                        numero_destinatario = remote_number
                    else:
                        numero_remetente = remote_number
                        numero_destinatario = instance

                    # Salva
                    msg_obj = WhatsappMessage.objects.create(
                        id_mensagem=id_msg,
                        instancia=instance,
                        de_mim=from_me,
                        numero_remetente=numero_remetente,
                        numero_destinatario=numero_destinatario,
                        texto=text or '[sem texto]',
                        tipo_mensagem=mtype,
                        url_media=media_url,
                        media_base64=media_base64 if mtype in ['image', 'audio'] else None,
                        timestamp=dt
                    )
                    
                    # Tenta linkar com Lead/Oportunidade
                    EvolutionService.identify_and_link_message(msg_obj)

                    # Encaminha para responsável do canal (se habilitado)
                    if not from_me and msg_obj.oportunidade_id:
                        try:
                            msg_obj.refresh_from_db()
                            opp = msg_obj.oportunidade
                            if opp and opp.canal and opp.canal.encaminhar_whatsapp_responsavel:
                                canal_obj = opp.canal
                                resp = canal_obj.responsavel
                                if resp and resp.telefone:
                                    resp_number = ''.join(filter(str.isdigit, resp.telefone))
                                    # Evita loop: não encaminha se o remetente é o próprio responsável
                                    if resp_number and resp_number not in (remote_number, f'55{remote_number}'):
                                        service_fwd = EvolutionService.for_canal(canal_obj)
                                        conta_nome = opp.conta.nome_empresa if opp.conta else ''
                                        contato_nome = opp.contato.nome if opp.contato else remote_number
                                        header = f"📩 *{contato_nome}*"
                                        if conta_nome:
                                            header += f" ({conta_nome})"
                                        fwd_text = f"{header}\n\n{text}"
                                        import threading as _threading
                                        def _forward(svc, num, txt):
                                            try:
                                                svc.send_text(num, txt)
                                            except Exception as fe:
                                                logger.error(f"[WEBHOOK] Erro ao encaminhar msg para responsável: {fe}")
                                        t = _threading.Thread(target=_forward, args=(service_fwd, resp_number, fwd_text))
                                        t.daemon = True
                                        t.start()
                        except Exception as fwd_err:
                            logger.error(f"[WEBHOOK] Erro ao verificar encaminhamento: {fwd_err}")

                    # Processamento assíncrono de mídia (imagens e áudios)
                    if needs_async_processing and mtype in ['image', 'audio']:
                        import threading
                        import time
                        
                        def process_media_async(msg_id, msg_key, instance_name, media_type):
                            """Processa mídia em background após delay para API estar pronta"""
                            try:
                                time.sleep(3)  # Aguarda API processar
                                
                                from django.db import connection
                                connection.close()  # Força nova conexão
                                
                                from .models import WhatsappMessage
                                from .services.evolution_api import EvolutionService
                                
                                msg = WhatsappMessage.objects.get(id=msg_id)
                                
                                # Obtém serviço do canal correto
                                service = EvolutionService(
                                    instance_name=instance_name,
                                    instance_token=None
                                )
                                
                                # Tenta pegar de configuração do canal via Oportunidade
                                if msg.oportunidade and msg.oportunidade.canal:
                                    service = EvolutionService(
                                        instance_name=msg.oportunidade.canal.evolution_instance_name or instance_name,
                                        instance_token=msg.oportunidade.canal.evolution_token
                                    )
                                
                                media_result = service.get_media_base64(msg_key)
                                
                                if media_result and media_result.get('base64'):
                                    if media_type == 'image':
                                        # Salva imagem
                                        mimetype = media_result.get('mimetype', 'image/jpeg')
                                        base64_data = media_result['base64']
                                        if not base64_data.startswith('data:'):
                                            base64_data = f"data:{mimetype};base64,{base64_data}"
                                        msg.media_base64 = base64_data
                                        msg.save(update_fields=['media_base64'])

                                    elif media_type == 'audio':
                                        # Salva base64 do áudio para reprodução futura
                                        mimetype = media_result.get('mimetype', 'audio/ogg; codecs=opus')
                                        audio_b64 = media_result['base64']
                                        if not audio_b64.startswith('data:'):
                                            audio_b64 = f"data:{mimetype};base64,{audio_b64}"
                                        msg.media_base64 = audio_b64

                                        # Transcreve áudio
                                        from .services.audio_transcription import transcribe_from_base64
                                        transcription = transcribe_from_base64(
                                            media_result['base64'],
                                            media_result.get('mimetype', '')
                                        )
                                        if transcription and transcription.get('text'):
                                            duration = transcription.get('duration', 0)
                                            msg.texto = f"🎤 [Áudio {int(duration)}s]: {transcription['text']}"
                                            msg.save(update_fields=['texto', 'media_base64'])
                                        else:
                                            msg.save(update_fields=['media_base64'])

                            except Exception as e:
                                logger.error(f"[ASYNC] Erro ao processar mídia {msg_id}: {e}")

                        # Dispara thread
                        msg_key_for_thread = {
                            'id': id_msg,
                            'remoteJid': remote_jid,
                            'fromMe': from_me
                        }
                        thread = threading.Thread(
                            target=process_media_async,
                            args=(msg_obj.id, msg_key_for_thread, instance, mtype)
                        )
                        thread.daemon = True
                        thread.start()

                except Exception as e:
                    logger.error(f"[WEBHOOK] Erro ao processar mensagem: {str(e)}")
                    continue

        return Response({'status': 'received'}, status=200)


class LogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet somente leitura para logs de auditoria.
    Permite consultar e filtrar logs do sistema.
    """
    serializer_class = LogSerializer
    permission_classes = [permissions.IsAuthenticated, HierarchyPermission]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['acao', 'modelo', 'usuario', 'objeto_id']
    search_fields = ['modelo', 'objeto_repr', 'observacao']
    ordering_fields = ['timestamp', 'acao', 'modelo', 'usuario']
    ordering = ['-timestamp']

    def get_queryset(self):
        """
        Retorna logs filtrados por hierarquia:
        - ADMIN: vê todos os logs
        - RESPONSAVEL: vê logs do seu canal (incluindo logs do sistema)
        - VENDEDOR: vê apenas seus próprios logs (incluindo logs do sistema relacionados)
        """
        user = self.request.user

        if user.perfil == 'ADMIN':
            queryset = Log.objects.all()
        elif user.perfil == 'RESPONSAVEL':
            # Responsável vê logs de todos os usuários do seu canal + logs do sistema
            canal_users = User.objects.filter(canal=user.canal)
            queryset = Log.objects.filter(
                models.Q(usuario__in=canal_users) | models.Q(usuario__isnull=True)
            )
        else:
            # Vendedor vê seus próprios logs + logs do sistema
            queryset = Log.objects.filter(
                models.Q(usuario=user) | models.Q(usuario__isnull=True)
            )

        return queryset.select_related('usuario')


class OrganogramaViewSet(viewsets.ViewSet):
    """
    ViewSet para gerar organograma da estrutura hierárquica.
    GET /api/organograma/ - Retorna SVG do organograma
    """
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request):
        """
        Gera o SVG do organograma hierárquico:
        Administrador > Canais (com Responsável) > Vendedores
        """
        # Busca todos os canais com seus responsáveis
        canais = list(Canal.objects.select_related('responsavel').all())
        
        # Busca todos os usuários agrupados por canal
        usuarios_por_canal = {}
        for user in User.objects.filter(perfil='VENDEDOR', is_active=True).select_related('canal'):
            if user.canal_id:
                if user.canal_id not in usuarios_por_canal:
                    usuarios_por_canal[user.canal_id] = []
                usuarios_por_canal[user.canal_id].append(user)
        
        # Configurações do layout
        box_width = 160
        box_height = 50
        box_spacing_x = 30
        box_spacing_y = 80
        start_y = 30
        
        # Cores para os canais
        cores_canais = ['#059669', '#0891b2', '#7c3aed', '#db2777', '#ea580c', '#65a30d']
        
        # Calcular largura total necessária para os canais
        total_canais = len(canais)
        canais_width = total_canais * box_width + (total_canais - 1) * box_spacing_x if total_canais > 0 else box_width
        
        # Calcular largura necessária para vendedores (máximo por canal)
        max_vendedores_por_canal = max([len(usuarios_por_canal.get(c.id, [])) for c in canais] + [0])
        
        # Largura do SVG
        svg_width = max(canais_width + 100, 800)
        
        # Posição X inicial dos canais (centralizado)
        canais_start_x = (svg_width - canais_width) / 2
        
        # Posição do Admin (centro superior)
        admin_x = svg_width / 2 - box_width / 2
        admin_y = start_y
        
        # Posição Y dos canais
        canais_y = admin_y + box_height + box_spacing_y
        
        # Calcular altura do SVG (considerando vendedores)
        vendedores_y = canais_y + box_height + box_spacing_y
        svg_height = vendedores_y + box_height + 50 if max_vendedores_por_canal > 0 else canais_y + box_height + 50
        
        # Construir o SVG
        svg_parts = [
            f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {svg_width} {svg_height}" class="organograma-svg">',
            '  <defs>',
            '    <marker id="arrowhead" markerWidth="10" markerHeight="7" refX="9" refY="3.5" orient="auto">',
            '      <polygon points="0 0, 10 3.5, 0 7" fill="#9ca3af"/>',
            '    </marker>',
            '    <filter id="shadow" x="-20%" y="-20%" width="140%" height="140%">',
            '      <feDropShadow dx="2" dy="2" stdDeviation="3" flood-opacity="0.15"/>',
            '    </filter>',
            '  </defs>',
            '',
            '  <!-- Administrador -->',
            f'  <rect x="{admin_x}" y="{admin_y}" width="{box_width}" height="{box_height}" rx="8" fill="#1e40af" filter="url(#shadow)"/>',
            f'  <text x="{admin_x + box_width/2}" y="{admin_y + box_height/2 + 5}" fill="white" text-anchor="middle" font-family="Arial, sans-serif" font-size="14" font-weight="bold">Administrador</text>',
        ]
        
        # Linhas do Admin para cada Canal e os Canais
        for i, canal in enumerate(canais):
            canal_x = canais_start_x + i * (box_width + box_spacing_x)
            canal_center_x = canal_x + box_width / 2
            admin_center_x = admin_x + box_width / 2
            
            # Linha do Admin para o Canal
            svg_parts.append(f'  <line x1="{admin_center_x}" y1="{admin_y + box_height}" x2="{canal_center_x}" y2="{canais_y}" stroke="#9ca3af" stroke-width="2" marker-end="url(#arrowhead)"/>')
            
            # Caixa do Canal
            cor = cores_canais[i % len(cores_canais)]
            responsavel_nome = canal.responsavel.get_full_name() if canal.responsavel else "Sem Responsavel"
            responsavel_nome = responsavel_nome if responsavel_nome.strip() else (canal.responsavel.username if canal.responsavel else "Sem Responsavel")
            # Escapa caracteres especiais para XML
            canal_nome = canal.nome.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            responsavel_nome = responsavel_nome.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            
            svg_parts.append(f'  <rect x="{canal_x}" y="{canais_y}" width="{box_width}" height="{box_height}" rx="8" fill="{cor}" filter="url(#shadow)"/>')
            svg_parts.append(f'  <text x="{canal_center_x}" y="{canais_y + 20}" fill="white" text-anchor="middle" font-family="Arial, sans-serif" font-size="12" font-weight="bold">{canal_nome}</text>')
            svg_parts.append(f'  <text x="{canal_center_x}" y="{canais_y + 36}" fill="white" text-anchor="middle" font-family="Arial, sans-serif" font-size="10" opacity="0.9">{responsavel_nome}</text>')
            
            # Vendedores deste canal
            vendedores = usuarios_por_canal.get(canal.id, [])
            num_vend = len(vendedores)
            
            if num_vend > 0:
                # Calcular posições dos vendedores (centralizados abaixo do canal)
                vend_box_width = 72  # 40% menor que antes
                vend_total_width = num_vend * vend_box_width + (num_vend - 1) * 15
                vend_start_x = canal_center_x - vend_total_width / 2
                
                for j, vend in enumerate(vendedores):
                    vend_x = vend_start_x + j * (vend_box_width + 15)
                    vend_center_x = vend_x + vend_box_width / 2
                    # Usar apenas o primeiro nome do vendedor
                    vend_nome_completo = vend.get_full_name() or vend.username
                    vend_nome = vend_nome_completo.split()[0] if vend_nome_completo else vend.username
                    vend_nome = vend_nome.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                    
                    # Linha do Canal para o Vendedor
                    svg_parts.append(f'  <line x1="{canal_center_x}" y1="{canais_y + box_height}" x2="{vend_center_x}" y2="{vendedores_y}" stroke="#9ca3af" stroke-width="2" marker-end="url(#arrowhead)"/>')
                    
                    # Caixa do Vendedor (menor)
                    svg_parts.append(f'  <rect x="{vend_x}" y="{vendedores_y}" width="{vend_box_width}" height="30" rx="4" fill="#f3f4f6" stroke="#d1d5db" stroke-width="1" filter="url(#shadow)"/>')
                    svg_parts.append(f'  <text x="{vend_center_x}" y="{vendedores_y + 19}" fill="#374151" text-anchor="middle" font-family="Arial, sans-serif" font-size="10">{vend_nome}</text>')
        
        svg_parts.append('</svg>')
        svg_code = '\n'.join(svg_parts)
        
        # Estatísticas
        total_canais = len(canais)
        total_vendedores = User.objects.filter(perfil='VENDEDOR', is_active=True).count()
        total_responsaveis = User.objects.filter(perfil='RESPONSAVEL', is_active=True).count()
        
        return Response({
            'svg': svg_code,
            'estatisticas': {
                'total_canais': total_canais,
                'total_responsaveis': total_responsaveis,
                'total_vendedores': total_vendedores
            }
        })

class TimelineViewSet(viewsets.ViewSet):
    """
    ViewSet unificado para a Timeline (Feed)
    Agrega Atividades, Mensagens do WhatsApp e Logs em uma única lista cronológica.
    Endpoint: /api/timeline/?model=oportunidade&id=1
    """
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request):
        model_name = request.query_params.get('model')
        model_id = request.query_params.get('id')
        
        if not model_name or not model_id:
            return Response(
                {'error': 'Parâmetros "model" e "id" são obrigatórios'}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        # 1. Identificar a entidade
        entity = None
        if model_name == 'oportunidade':
            entity = Oportunidade.objects.filter(id=model_id).first()
        elif model_name == 'contato':
            entity = Contato.objects.filter(id=model_id).first()
        elif model_name == 'conta':
            entity = Conta.objects.filter(id=model_id).first()
        
        if not entity:
            return Response({'error': 'Objeto não encontrado'}, status=status.HTTP_404_NOT_FOUND)

        # 2. Buscar Dados
        timeline_items = []
        related_oportunidade_ids = []
        related_conta_id = None
        phone_lookup_values = set()

        def add_phone_variants(raw_phone):
            digits = re.sub(r'\D', '', str(raw_phone or ''))
            if not digits:
                return

            candidates = {digits}
            if digits.startswith('55') and len(digits) > 2:
                candidates.add(digits[2:])
            elif len(digits) <= 11:
                candidates.add(f"55{digits}")

            for number in candidates:
                if not number:
                    continue
                phone_lookup_values.add(number)
                phone_lookup_values.add(f"{number}@s.whatsapp.net")

        if model_name == 'contato':
            related_conta_id = entity.conta_id
            related_oportunidade_ids = list(
                Oportunidade.objects.filter(
                    Q(contato_principal=entity) | Q(contatos=entity)
                ).values_list('id', flat=True).distinct()
            )

            add_phone_variants(entity.telefone)
            add_phone_variants(entity.celular)
            for phone in entity.telefones.values_list('numero', flat=True):
                add_phone_variants(phone)
        elif model_name == 'conta':
            related_oportunidade_ids = list(
                Oportunidade.objects.filter(
                    Q(conta=entity) | Q(empresas=entity)
                ).values_list('id', flat=True).distinct()
            )
            add_phone_variants(entity.telefone_principal)

        # A. Atividades (Padrão Django ContentType)
        # Precisamos descobrir o ContentType do modelo
        from django.contrib.contenttypes.models import ContentType
        ct = ContentType.objects.get_for_model(entity)
        atividades_filter = Q(content_type=ct, object_id=entity.id)

        if model_name == 'contato':
            if related_conta_id:
                ct_conta = ContentType.objects.get_for_model(Conta)
                atividades_filter |= Q(content_type=ct_conta, object_id=related_conta_id)
            if related_oportunidade_ids:
                ct_oportunidade = ContentType.objects.get_for_model(Oportunidade)
                atividades_filter |= Q(content_type=ct_oportunidade, object_id__in=related_oportunidade_ids)
        elif model_name == 'conta' and related_oportunidade_ids:
            ct_oportunidade = ContentType.objects.get_for_model(Oportunidade)
            atividades_filter |= Q(content_type=ct_oportunidade, object_id__in=related_oportunidade_ids)

        atividades = Atividade.objects.filter(atividades_filter).select_related('proprietario')
        
        for item in atividades:
            timeline_items.append({
                'id': f"atividade_{item.id}",
                'db_id': item.id,
                'type': 'atividade',
                'subtype': item.tipo, # TAREFA, LIGACAO, NOTA, etc
                'timestamp': item.data_criacao,
                'author': item.proprietario.get_full_name() if item.proprietario else 'Sistema',
                'content': item.descricao or item.titulo,
                'title': item.titulo,
                'status': item.status,
                'data': AtividadeSerializer(item).data
            })

        # B. WhatsApp Messages
        # Lógica de vínculo varia por modelo
        msgs = WhatsappMessage.objects.none()
        if model_name == 'oportunidade':
            # Mensagens vinculadas diretamente à oportunidade
            # OU mensagens do contato principal da oportunidade (opcional, pode ser complexo filtrar data)
            msgs = WhatsappMessage.objects.filter(oportunidade=entity)
        elif model_name in ['contato', 'conta']:
            msg_filter = Q()
            has_msg_filters = False

            if related_oportunidade_ids:
                msg_filter |= Q(oportunidade_id__in=related_oportunidade_ids)
                has_msg_filters = True

            if phone_lookup_values:
                msg_filter |= Q(numero_remetente__in=phone_lookup_values) | Q(numero_destinatario__in=phone_lookup_values)
                has_msg_filters = True

            if has_msg_filters:
                msgs = WhatsappMessage.objects.filter(msg_filter).distinct()
        
        for item in msgs:
            timeline_items.append({
                'id': f"whatsapp_{item.id}",
                'db_id': item.id,
                'type': 'whatsapp',
                'subtype': item.tipo_mensagem, # text, image, audio
                'timestamp': item.timestamp,
                'author': 'Eu' if item.de_mim else (item.numero_remetente or 'Cliente'),
                'direction': 'outbound' if item.de_mim else 'inbound',
                'content': item.texto or f"[{item.tipo_mensagem or 'midia'}]",
                'status': 'read' if item.lida else 'delivered',
                'data': WhatsappMessageSerializer(item).data
            })

        # C. Logs (Audit)
        log_filter = Q(modelo=entity.__class__.__name__, objeto_id=entity.id)

        if model_name == 'contato':
            if related_conta_id:
                log_filter |= Q(modelo='Conta', objeto_id=related_conta_id)
            if related_oportunidade_ids:
                log_filter |= Q(modelo='Oportunidade', objeto_id__in=related_oportunidade_ids)
        elif model_name == 'conta' and related_oportunidade_ids:
            log_filter |= Q(modelo='Oportunidade', objeto_id__in=related_oportunidade_ids)

        logs = Log.objects.filter(log_filter)
        for item in logs:
             timeline_items.append({
                'id': f"log_{item.id}",
                'db_id': item.id,
                'type': 'log',
                'subtype': item.acao,
                'timestamp': item.timestamp,
                'author': item.usuario.get_full_name() if item.usuario else 'Sistema',
                'content': f"{item.get_acao_display()}: {item.observacao or ''}",
                'data': LogSerializer(item).data
            })

        # 3. Ordenação (Do mais recente para o mais antigo)
        timeline_items.sort(key=lambda x: x['timestamp'], reverse=True)

        # 4. Paginação Manual
        page = int(request.query_params.get('page', 1))
        page_size = 20
        start = (page - 1) * page_size
        end = start + page_size
        
        paginated_items = timeline_items[start:end]
        
        return Response({
            'count': len(timeline_items),
            'results': paginated_items,
            'page': page,
            'pages': max(1, (len(timeline_items) + page_size - 1) // page_size)
        })
